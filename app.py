import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import sqlite3
import os
import tempfile
from data_processor import DataProcessor
from cf_analyzer import CashFlowAnalyzer
from datetime import datetime
from profitability_analyzer import ProfitabilityAnalyzer, analyze_profitability_from_db
from cfo_advisor import CFOAdvisor
from profitability_analysis_ui import show_profitability_analysis_page

# ==================== 完全ライトモード設定 ====================
st.markdown("""
<style>
    /* 全体の背景を白に統一 */
    .stApp {
        background-color: #FFFFFF !important;
        color: #262730 !important;
    }
    
    /* メインコンテンツエリア */
    .main {
        background-color: #FFFFFF !important;
    }
    
    .main .block-container {
        background-color: #FFFFFF !important;
    }
    
    /* すべての要素の背景 */
    .element-container {
        background-color: transparent !important;
    }
    
    /* カード・パネル */
    .stAlert, .stInfo, .stWarning, .stError, .stSuccess {
        background-color: #F0F2F6 !important;
        color: #262730 !important;
    }
    
    /* データフレーム */
    .dataframe {
        background-color: #FFFFFF !important;
        color: #262730 !important;
    }
    
    /* エクスパンダー */
    .streamlit-expanderHeader {
        background-color: #F0F2F6 !important;
        color: #262730 !important;
    }
    
    .streamlit-expanderContent {
        background-color: #FFFFFF !important;
        color: #262730 !important;
    }
    
    /* メトリクス */
    .stMetric {
        background-color: #F0F2F6 !important;
    }
    
    /* タブ */
    .stTabs [data-baseweb="tab-list"] {
        background-color: #F0F2F6 !important;
    }
    
    .stTabs [data-baseweb="tab"] {
        background-color: #FFFFFF !important;
        color: #262730 !important;
    }
    
    /* テキスト入力 */
    .stTextInput > div > div {
        background-color: #FFFFFF !important;
        color: #262730 !important;
    }
    
    /* セレクトボックス */
    .stSelectbox > div > div {
        background-color: #FFFFFF !important;
        color: #262730 !important;
    }
    
    /* コードブロック */
    .stCodeBlock {
        background-color: #F0F2F6 !important;
    }
    
    code {
        background-color: #F0F2F6 !important;
        color: #262730 !important;
    }
    
    /* マークダウン */
    .stMarkdown {
        color: #262730 !important;
    }
    
    /* ボタン */
    .stButton > button {
        background-color: #FFFFFF !important;
        color: #262730 !important;
        border: 1px solid #E0E0E0 !important;
    }
    
    .stButton > button:hover {
        background-color: #F0F2F6 !important;
        border-color: #1f77b4 !important;
    }
    
    /* プライマリボタン */
    .stButton > button[kind="primary"] {
        background-color: #1f77b4 !important;
        color: #FFFFFF !important;
    }
    
    /* チェックボックス・ラジオボタン */
    .stCheckbox, .stRadio {
        color: #262730 !important;
    }
    
    /* サイドバーのテキスト */
    section[data-testid="stSidebar"] * {
        color: #262730 !important;
    }
</style>
""", unsafe_allow_html=True)



# ==================== 可視化関数 ====================

def create_bs_sankey(bs_data):
    """BS Sankey図作成 - 左:資産、右:負債・純資産"""
    try:
        if bs_data is None or bs_data.empty:
            return None
        
        # データ準備
        source = []
        target = []
        value = []
        labels = []
        colors = []
        
        label_dict = {}
        label_counter = 0
        
        def add_label(label, color="#ADD8E6"):
            nonlocal label_counter
            if label not in label_dict:
                label_dict[label] = label_counter
                labels.append(label)
                colors.append(color)
                label_counter += 1
            return label_dict[label]
        
        # 集計用辞書
        assets_current = {}  # 流動資産
        assets_fixed = {}    # 固定資産
        liabilities_current = {}  # 流動負債
        liabilities_fixed = {}    # 固定負債
        equity = {}          # 純資産
        
        # データ分類
        for _, row in bs_data.iterrows():
            item = str(row.get('項目名', ''))
            amount = abs(float(row.get('金額', 0)))
            
            if amount == 0:
                continue
            
            # 流動資産
            if any(x in item for x in ['現金', '預金', '当座', '普通', '定期', '外貨']):
                if '合計' not in item:
                    assets_current[item] = amount
            elif any(x in item for x in ['売掛', '売上債権']):
                if '合計' not in item:
                    assets_current[item] = amount
            elif any(x in item for x in ['棚卸', '商品', '貯蔵']):
                if '合計' not in item:
                    assets_current[item] = amount
            elif any(x in item for x in ['立替', '前払', '未収', '仮払']):
                if '合計' not in item and '法人税' not in item:
                    assets_current[item] = amount
            
            # 固定資産
            elif any(x in item for x in ['有形固定', '附属設備', '車両']):
                if '合計' not in item and '計' not in item:
                    assets_fixed[item] = amount
            elif any(x in item for x in ['投資有価証券', '関係会社株式', '出資金', '敷金', '差入保証', '長期貸付', '保険積立']):
                assets_fixed[item] = amount
            
            # 流動負債
            elif any(x in item for x in ['買掛', '仕入債務', '短期借入', '未払', '預り', '仮受']):
                if '合計' not in item:
                    liabilities_current[item] = amount
            
            # 固定負債
            elif any(x in item for x in ['長期借入']):
                if '合計' not in item:
                    liabilities_fixed[item] = amount
            
            # 純資産
            elif any(x in item for x in ['資本金', '利益準備金', '繰越利益剰余金']):
                equity[item] = amount
        
        # ノード定義（左から右へ）
        # レベル0: 総資産
        total_assets_idx = add_label("資産合計", "#4A90E2")
        
        # レベル1: 流動資産・固定資産
        current_assets_idx = add_label("流動資産", "#7CB9E8")
        fixed_assets_idx = add_label("固定資産", "#6CA0DC")
        
        # レベル2: 資産詳細（左側）
        current_asset_indices = {}
        for item, amt in assets_current.items():
            current_asset_indices[item] = add_label(item, "#B0D4F1")
        
        fixed_asset_indices = {}
        for item, amt in assets_fixed.items():
            fixed_asset_indices[item] = add_label(item, "#90C4E8")
        
        # レベル3: 負債・純資産合計
        total_liab_equity_idx = add_label("負債・純資産", "#E89090")
        
        # レベル4: 流動負債・固定負債・純資産
        current_liab_idx = add_label("流動負債", "#F0A0A0")
        fixed_liab_idx = add_label("固定負債", "#E88080")
        equity_idx = add_label("純資産", "#90E890")
        
        # レベル5: 負債・純資産詳細（右側）
        current_liab_indices = {}
        for item, amt in liabilities_current.items():
            current_liab_indices[item] = add_label(item, "#F5C0C0")
        
        fixed_liab_indices = {}
        for item, amt in liabilities_fixed.items():
            fixed_liab_indices[item] = add_label(item, "#F0B0B0")
        
        equity_indices = {}
        for item, amt in equity.items():
            equity_indices[item] = add_label(item, "#B0F0B0")
        
        # リンク作成
        # 資産側（左）
        total_current_assets = sum(assets_current.values())
        total_fixed_assets = sum(assets_fixed.values())
        
        if total_current_assets > 0:
            source.append(total_assets_idx)
            target.append(current_assets_idx)
            value.append(total_current_assets)
            
            for item, amt in assets_current.items():
                source.append(current_assets_idx)
                target.append(current_asset_indices[item])
                value.append(amt)
        
        if total_fixed_assets > 0:
            source.append(total_assets_idx)
            target.append(fixed_assets_idx)
            value.append(total_fixed_assets)
            
            for item, amt in assets_fixed.items():
                source.append(fixed_assets_idx)
                target.append(fixed_asset_indices[item])
                value.append(amt)
        
        # 負債・純資産側（右）
        total_current_liab = sum(liabilities_current.values())
        total_fixed_liab = sum(liabilities_fixed.values())
        total_equity = sum(equity.values())
        total_liab_equity = total_current_liab + total_fixed_liab + total_equity
        
        if total_liab_equity > 0:
            if total_current_liab > 0:
                source.append(current_liab_idx)
                target.append(total_liab_equity_idx)
                value.append(total_current_liab)
                
                for item, amt in liabilities_current.items():
                    source.append(current_liab_indices[item])
                    target.append(current_liab_idx)
                    value.append(amt)
            
            if total_fixed_liab > 0:
                source.append(fixed_liab_idx)
                target.append(total_liab_equity_idx)
                value.append(total_fixed_liab)
                
                for item, amt in liabilities_fixed.items():
                    source.append(fixed_liab_indices[item])
                    target.append(fixed_liab_idx)
                    value.append(amt)
            
            if total_equity > 0:
                source.append(equity_idx)
                target.append(total_liab_equity_idx)
                value.append(total_equity)
                
                for item, amt in equity.items():
                    source.append(equity_indices[item])
                    target.append(equity_idx)
                    value.append(amt)
        
        if not source:
            return None
        
        # Sankey図作成
        fig = go.Figure(data=[go.Sankey(
            arrangement='snap',
            node=dict(
                pad=20,
                thickness=25,
                line=dict(color="white", width=2),
                label=labels,
                color=colors,
                customdata=[f"¥{sum([value[i] for i in range(len(value)) if target[i] == idx or source[i] == idx]):,.0f}" for idx in range(len(labels))],
                hovertemplate='%{label}<br>¥%{customdata}<extra></extra>'
            ),
            link=dict(
                source=source,
                target=target,
                value=value,
                color="rgba(200,200,200,0.3)",
                hovertemplate='%{source.label} → %{target.label}<br>¥%{value:,.0f}<extra></extra>'
            )
        )])
        
        fig.update_layout(
            title={
                'text': "貸借対照表の構造（左:資産、右:負債・純資産）",
                'x': 0.5,
                'xanchor': 'center'
            },
            font=dict(size=11, family="sans-serif"),
            height=700,
            paper_bgcolor='rgba(0,0,0,0)',
            plot_bgcolor='rgba(0,0,0,0)'
        )
        
        return fig
    except Exception as e:
        import sys
        import traceback
        sys.stderr.write(f"BS Sankey error: {e}\n")
        sys.stderr.write(traceback.format_exc())
        return None


def create_pl_waterfall(pl_data):
    """PLウォーターフォール（横持ち対応）"""
    try:
        # データが横持ちの場合（月が列）
        if '項目名' in pl_data.columns:
            # 合計列を使用
            if '合計' in pl_data.columns:
                sales = cogs = sg_expense = operating_profit = 0
                
                for _, row in pl_data.iterrows():
                    item = str(row.get('項目名', ''))
                    amount = float(row.get('合計', 0))
                    
                    if item == '売上高':
                        sales = amount
                    elif item == '売上原価':
                        cogs = amount
                    elif '販売' in item and '管理費' in item:
                        sg_expense = amount
                    elif item == '営業損益金額':
                        operating_profit = amount
            else:
                # 最新月の列を使用
                month_cols = [col for col in pl_data.columns if '-' in str(col) or col.isdigit()]
                if month_cols:
                    latest_month = month_cols[-1]
                    
                    for _, row in pl_data.iterrows():
                        item = str(row.get('項目名', ''))
                        amount = float(row.get(latest_month, 0))
                        
                        if item == '売上高':
                            sales = amount
                        elif item == '売上原価':
                            cogs = amount
                        elif '販売' in item and '管理費' in item:
                            sg_expense = amount
                        elif item == '営業損益金額':
                            operating_profit = amount
                else:
                    return None
        else:
            # 縦持ち形式
            sales = cogs = sg_expense = operating_profit = 0
            for _, row in pl_data.iterrows():
                item = str(row.get('項目名', ''))
                amount = float(row.get('金額', 0))
                if item == '売上高':
                    sales = amount
                elif item == '売上原価':
                    cogs = amount
                elif '販売' in item and '管理費' in item:
                    sg_expense = amount
                elif item == '営業損益金額':
                    operating_profit = amount
        
        # グラフ作成
        gross_profit = sales - cogs
        x_labels = ['売上高', '売上原価', '売上総利益', '販管費', '営業利益']
        y_values = [sales, -cogs, gross_profit, -sg_expense, operating_profit]
        
        fig = go.Figure(go.Waterfall(
            x=x_labels, 
            y=y_values,
            measure=['absolute', 'relative', 'total', 'relative', 'total'],
            text=[f"¥{abs(v):,.0f}" for v in y_values],
            textposition='outside',
            increasing={"marker": {"color": "#2ecc71"}},
            decreasing={"marker": {"color": "#e74c3c"}},
            totals={"marker": {"color": "#3498db"}}
        ))
        
        fig.update_layout(
            title="損益の流れ", 
            showlegend=False, 
            height=500,
            yaxis_title="金額（円）"
        )
        return fig
    except Exception as e:
        import sys
        sys.stderr.write(f"Waterfall error: {e}\n")
        import traceback
        sys.stderr.write(traceback.format_exc())
        return None


# AI Forecast import
try:
    from advanced_forecast_engine import get_advanced_forecast_engine
    from advanced_forecast_ui import show_advanced_forecast_page
    ADVANCED_FORECAST_AVAILABLE = True
except ImportError:
    ADVANCED_FORECAST_AVAILABLE = False


# ページ設定 - 完全ライトモード
st.set_page_config(
    page_title="財務予測シミュレーター",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Streamlit標準テーマを強制的にライトモードに
st.markdown("""
<script>
    // ライトモードを強制
    var theme = window.parent.document.querySelector('[data-testid="stAppViewContainer"]');
    if (theme) {
        theme.style.backgroundColor = "#fafbfc";
    }
</script>
""", unsafe_allow_html=True)

# カスタムCSS - Manageboard風デザイン（実際のUIに準拠）
st.markdown("""
<style>
    /* 全体背景 - Manageboardの明るいグレー */
    .main {
        padding: 0rem 1rem;
        background-color: #fafbfc;
    }
    
    /* タイトル - よりシンプルに */
    h1 {
        color: #2c3e50;
        font-weight: 600;
        margin-bottom: 1.5rem;
        font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
        font-size: 1.75rem;
    }
    
    h2 {
        color: #34495e;
        font-weight: 600;
        margin-top: 2rem;
        margin-bottom: 1rem;
        font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
        font-size: 1.25rem;
    }
    
    h3 {
        color: #5a6c7d;
        font-weight: 600;
        font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
        font-size: 1rem;
    }
    
    /* 金額カード - Manageboardスタイル */
    .amount-card {
        background: #ffffff;
        padding: 1.25rem 1.5rem;
        border-radius: 6px;
        border: 1px solid #e1e8ed;
        box-shadow: 0 1px 3px rgba(0, 0, 0, 0.04);
        margin-bottom: 1rem;
        transition: all 0.2s ease;
    }
    
    .amount-card:hover {
        box-shadow: 0 2px 8px rgba(0, 0, 0, 0.08);
    }
    
    .amount-card-label {
        font-size: 0.75rem;
        color: #8a9ba8;
        font-weight: 600;
        margin-bottom: 0.5rem;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    
    .amount-card-value {
        font-size: 1.75rem;
        font-weight: 700;
        color: #2c3e50;
        margin-bottom: 0.5rem;
        font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
    }
    
    .amount-card-sub {
        font-size: 0.8rem;
        color: #8a9ba8;
        margin-top: 0.3rem;
    }
    
    /* サマリーカード - よりフラットに */
    .summary-card {
        background: #ffffff;
        padding: 1.25rem 1.5rem;
        border-radius: 6px;
        border: 1px solid #e1e8ed;
        box-shadow: 0 1px 3px rgba(0, 0, 0, 0.04);
        margin-bottom: 1rem;
    }
    
    .summary-card-blue {
        background: #ffffff;
        padding: 1.25rem 1.5rem;
        border-radius: 6px;
        border-left: 3px solid #3b82f6;
        box-shadow: 0 1px 3px rgba(0, 0, 0, 0.04);
        margin-bottom: 1rem;
    }
    
    .summary-card-green {
        background: #ffffff;
        padding: 1.25rem 1.5rem;
        border-radius: 6px;
        border-left: 3px solid #10b981;
        box-shadow: 0 1px 3px rgba(0, 0, 0, 0.04);
        margin-bottom: 1rem;
    }
    
    .summary-card-orange {
        background: #ffffff;
        padding: 1.25rem 1.5rem;
        border-radius: 6px;
        border-left: 3px solid #f59e0b;
        box-shadow: 0 1px 3px rgba(0, 0, 0, 0.04);
        margin-bottom: 1rem;
    }
    
    .summary-card-purple {
        background: #ffffff;
        padding: 1.25rem 1.5rem;
        border-radius: 6px;
        border-left: 3px solid #8b5cf6;
        box-shadow: 0 1px 3px rgba(0, 0, 0, 0.04);
        margin-bottom: 1rem;
    }
    
    .card-title {
        font-size: 0.75rem;
        font-weight: 600;
        margin-bottom: 0.5rem;
        color: #8a9ba8;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    
    .card-value {
        font-size: 1.75rem;
        font-weight: 700;
        margin-bottom: 0.3rem;
        font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
        color: #2c3e50;
    }
    
    .card-subtitle {
        font-size: 0.8rem;
        color: #8a9ba8;
        font-weight: 400;
    }
    
    /* インフォボックス */
    .info-box {
        background-color: #f0f9ff;
        border-left: 3px solid #3b82f6;
        padding: 0.875rem 1.25rem;
        border-radius: 6px;
        margin-bottom: 1.5rem;
        font-size: 0.875rem;
        color: #1e40af;
    }
    
    .warning-box {
        background-color: #fffbeb;
        border-left: 3px solid #f59e0b;
        padding: 0.875rem 1.25rem;
        border-radius: 6px;
        margin-bottom: 1.5rem;
        font-size: 0.875rem;
        color: #92400e;
    }
    
    .success-box {
        background-color: #f0fdf4;
        border-left: 3px solid #10b981;
        padding: 0.875rem 1.25rem;
        border-radius: 6px;
        margin-bottom: 1.5rem;
        font-size: 0.875rem;
        color: #065f46;
    }
    
    /* テーブルスタイル */
    .dataframe {
        border: 1px solid #e1e8ed !important;
        border-radius: 6px;
        overflow: hidden;
        box-shadow: 0 1px 3px rgba(0, 0, 0, 0.04);
        font-size: 0.875rem;
    }
    
    .dataframe thead tr th {
        background-color: #f8fafc !important;
        color: #475569 !important;
        font-weight: 600 !important;
        padding: 12px 16px !important;
        border-bottom: 2px solid #e1e8ed !important;
        text-transform: uppercase;
        font-size: 0.75rem;
        letter-spacing: 0.5px;
    }
    
    .dataframe tbody tr {
        border-bottom: 1px solid #f1f5f9 !important;
    }
    
    .dataframe tbody tr:hover {
        background-color: #f8fafc !important;
    }
    
    .dataframe tbody td {
        padding: 12px 16px !important;
        color: #334155;
    }
    
    /* マネージボード風スタイル */
    .kpi-card {
        border-radius: 16px;
        padding: 24px;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
        color: white;
        position: relative;
        overflow: hidden;
        transition: transform 0.2s ease, box-shadow 0.2s ease;
        margin-bottom: 24px;
    }
    
    .kpi-card:hover {
        transform: translateY(-4px);
        box-shadow: 0 10px 15px -3px rgba(0, 0, 0, 0.15);
    }
    
    .kpi-card-title {
        font-size: 0.875rem;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 1px;
        opacity: 0.9;
        margin-bottom: 12px;
    }
    
    .kpi-card-value {
        font-size: 2.5rem;
        font-weight: 700;
        font-variant-numeric: tabular-nums;
        margin-bottom: 8px;
        line-height: 1.1;
    }
    
    .kpi-card-subtitle {
        font-size: 0.875rem;
        opacity: 0.8;
        margin-bottom: 12px;
    }
    
    .kpi-card-trend {
        display: inline-block;
        background: rgba(255, 255, 255, 0.2);
        border-radius: 20px;
        padding: 6px 12px;
        font-size: 0.875rem;
        font-weight: 600;
    }
    
    .kpi-card-decoration {
        position: absolute;
        right: -30px;
        bottom: -30px;
        width: 150px;
        height: 150px;
        background: rgba(255, 255, 255, 0.1);
        border-radius: 50%;
    }
    
    .dashboard-header {
        background: linear-gradient(90deg, #1F2937 0%, #374151 100%);
        padding: 24px 32px;
        border-radius: 12px;
        margin-bottom: 32px;
        color: white;
    }
    
    .dashboard-title {
        margin: 0;
        font-size: 2rem;
        font-weight: 700;
    }
    
    .dashboard-subtitle {
        margin: 8px 0 0 0;
        opacity: 0.8;
        font-size: 0.875rem;
    }
    
    .section-card {
        background: white;
        border-radius: 16px;
        padding: 24px;
        box-shadow: 0 1px 3px rgba(0, 0, 0, 0.1);
        margin-bottom: 24px;
    }
    
    .section-title {
        margin: 0 0 16px 0;
        font-size: 1.25rem;
        font-weight: 600;
        color: #1F2937;
    }
    
    .metric-row {
        display: flex;
        justify-content: space-between;
        align-items: center;
        padding: 12px 0;
        border-bottom: 1px solid #F3F4F6;
    }
    
    .metric-row:last-child {
        border-bottom: none;
    }
    
    .metric-name {
        font-size: 0.875rem;
        color: #6B7280;
        font-weight: 500;
    }
    
    .metric-value {
        font-size: 1.125rem;
        font-weight: 600;
        color: #1F2937;
        font-variant-numeric: tabular-nums;
    }
    
    .metric-change {
        font-size: 0.875rem;
        font-weight: 600;
        margin-left: 8px;
    }
    
    .metric-up {
        color: #10B981;
    }
    
    .metric-down {
        color: #EF4444;
    }
    
    @keyframes fadeInUp {
        from {
            opacity: 0;
            transform: translateY(20px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }
    
    .fade-in {
        animation: fadeInUp 0.6s ease-out;
    }
    
    @media (max-width: 768px) {
        .kpi-card-value {
            font-size: 2rem;
        }
        .dashboard-title {
            font-size: 1.5rem;
        }
    }
    
    /* ボタンスタイル */
    .stButton > button {
        border-radius: 6px;
        font-weight: 500;
        font-size: 0.875rem;
        padding: 0.5rem 1rem;
        transition: all 0.2s ease;
        border: 1px solid #e1e8ed;
        background-color: #ffffff;
        color: #475569;
    }
    
    .stButton > button:hover {
        background-color: #f8fafc;
        border-color: #cbd5e1;
        transform: translateY(-1px);
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.08);
    }
    
    .stButton > button[kind="primary"] {
        background-color: #3b82f6;
        color: white;
        border: none;
    }
    
    .stButton > button[kind="primary"]:hover {
        background-color: #2563eb;
    }
    
    /* サイドバー - 完全ライトモード */
    [data-testid="stSidebar"] {
        background-color: #f8fafc !important;
        border-right: 1px solid #e1e8ed;
    }
    
    [data-testid="stSidebar"] * {
        color: #1e293b !important;
    }
    
    [data-testid="stSidebar"] .stMarkdown {
        color: #1e293b !important;
    }
    
    [data-testid="stSidebar"] p {
        color: #1e293b !important;
    }
    
    /* セクション見出し（サイドバー内） */
    [data-testid="stSidebar"] h3 {
        color: #0f172a !important;
        font-size: 0.7rem;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: 1.2px;
        margin-top: 1.5rem;
        margin-bottom: 0.75rem;
        padding-left: 0.5rem;
        background-color: transparent !important;
    }
    
    /* サイドバーのボタン - 見やすく改善 */
    [data-testid="stSidebar"] .stButton > button {
        width: 100%;
        text-align: left;
        padding: 0.75rem 1rem;
        margin-bottom: 0.5rem;
        background-color: #ffffff !important;  /* 白背景で目立たせる */
        border: 1px solid #cbd5e1 !important;  /* 枠線を追加 */
        color: #1e293b !important;
        font-weight: 600;  /* 太字に */
        font-size: 0.9rem;  /* 少し大きく */
        border-radius: 8px;  /* 角丸を大きく */
        box-shadow: 0 1px 2px rgba(0, 0, 0, 0.05);  /* 軽い影 */
        transition: all 0.2s ease;
    }
    
    [data-testid="stSidebar"] .stButton > button:hover {
        background-color: #3b82f6 !important;  /* ホバー時は青 */
        color: #ffffff !important;  /* ホバー時は白文字 */
        border-color: #3b82f6 !important;
        transform: translateX(4px);  /* 右に少し移動 */
        box-shadow: 0 4px 6px rgba(59, 130, 246, 0.2);  /* 青い影 */
    }
    
    /* セレクトボックスの文字色 */
    .stSelectbox label {
        color: #1e293b !important;
    }
    
    .stSelectbox > div > div {
        background-color: #ffffff !important;
        color: #1e293b !important;
    }
    
    .stSelectbox [data-baseweb="select"] {
        background-color: #ffffff !important;
    }
    
    .stSelectbox [data-baseweb="select"] > div {
        color: #1e293b !important;
        background-color: #ffffff !important;
    }
    
    /* タブスタイル */
    .stTabs [data-baseweb="tab-list"] {
        gap: 4px;
        background-color: transparent;
        padding: 0;
        border-bottom: 2px solid #e1e8ed;
    }
    
    .stTabs [data-baseweb="tab"] {
        font-weight: 500;
        color: #64748b;
        border-radius: 0;
        padding: 0.75rem 1.5rem;
        border-bottom: 2px solid transparent;
        margin-bottom: -2px;
        background-color: transparent;
        font-size: 0.875rem;
    }
    
    .stTabs [aria-selected="true"] {
        background-color: transparent;
        color: #3b82f6;
        border-bottom: 2px solid #3b82f6;
        font-weight: 600;
    }
    
    /* セレクトボックス */
    .stSelectbox > div > div {
        background-color: #ffffff;
        border: 1px solid #e1e8ed;
        border-radius: 6px;
        font-size: 0.875rem;
    }
    
    /* データエディタ */
    [data-testid="stDataFrameResizable"] {
        border: 1px solid #e1e8ed;
        border-radius: 6px;
        box-shadow: 0 1px 3px rgba(0, 0, 0, 0.04);
    }
    
    /* メトリクス */
    [data-testid="stMetricValue"] {
        font-size: 1.75rem;
        font-weight: 700;
        color: #2c3e50;
    }
    
    [data-testid="stMetricLabel"] {
        font-size: 0.75rem;
        color: #8a9ba8;
        font-weight: 600;
        text-transform: uppercase;
    }
    
    /* セパレーター */
    hr {
        border: none;
        border-top: 1px solid #e1e8ed;
        margin: 1.5rem 0;
    }
</style>
""", unsafe_allow_html=True)

# 初期化
if 'processor' not in st.session_state:
    st.session_state.processor = DataProcessor()
processor = st.session_state.processor

# CFアナライザーの初期化
if 'cf_analyzer' not in st.session_state:
    st.session_state.cf_analyzer = CashFlowAnalyzer(processor)
cf_analyzer = st.session_state.cf_analyzer

# キャッシュ付きデータ読み込み関数（高速化）
@st.cache_data(ttl=600)  # 10分間キャッシュ（パフォーマンス改善）
def load_actual_data_cached(period_id, _processor):
    """実績データをキャッシュ付きで読み込み"""
    return _processor.load_actual_data(period_id)

@st.cache_data(ttl=600)  # 10分間キャッシュ（パフォーマンス改善）
def load_forecast_data_cached(period_id, scenario, _processor):
    """予測データをキャッシュ付きで読み込み"""
    return _processor.load_forecast_data(period_id, scenario)

@st.cache_data(ttl=600)  # 10分間キャッシュ（パフォーマンス改善）
def load_sub_accounts_cached(period_id, scenario, _processor):
    """補助科目データをキャッシュ付きで読み込み"""
    return _processor.load_sub_accounts(period_id, scenario)

@st.cache_data(ttl=3600)  # 1時間キャッシュ（マスタデータ）
def get_companies_cached(_processor):
    """会社一覧をキャッシュ付きで取得"""
    return _processor.get_companies()

@st.cache_data(ttl=3600)  # 1時間キャッシュ（マスタデータ）
def get_company_periods_cached(comp_id, _processor):
    """会計期間一覧をキャッシュ付きで取得"""
    return _processor.get_company_periods(comp_id)

@st.cache_data(ttl=3600)  # 1時間キャッシュ（マスタデータ）
def get_fiscal_months_cached(comp_id, period_id, _processor):
    """会計月一覧をキャッシュ付きで取得"""
    return _processor.get_fiscal_months(comp_id, period_id)

# ヘルパー関数: 安全なint変換
def safe_int(value):
    """NaN/None対応の安全なint変換"""
    try:
        if pd.isna(value) or value is None:
            return 0
        return int(float(value))
    except (ValueError, TypeError):
        return 0


def safe_float(value):
    """NaN/None対応の安全なfloat変換"""
    try:
        if pd.isna(value) or value is None:
            return 0.0
        return float(value)
    except (ValueError, TypeError):
        return 0.0


# データ形式変換関数
def convert_wide_to_long(df):
    """横持ち→縦持ち変換"""
    try:
        # 項目名カラムを探す
        item_col = None
        for col in df.columns:
            if col in ['項目名', 'item_name', '科目', 'account_name', 'account']:
                item_col = col
                break
        
        if not item_col:
            return df
        
        # 他のカラムが日付形式かチェック
        other_cols = [c for c in df.columns if c != item_col]
        if not other_cols:
            return df
        
        sample = str(other_cols[0])
        if not ('-' in sample or '/' in sample):
            return df
        
        # 縦持ちに変換
        long_df = df.melt(
            id_vars=[item_col],
            value_vars=other_cols,
            var_name='period',
            value_name='amount'
        )
        
        # 月番号抽出
        def get_month(s):
            try:
                s = str(s)
                if '-' in s:
                    return int(s.split('-')[-1])
                if '/' in s:
                    return int(s.split('/')[-1])
                return 0
            except:
                return 0
        
        long_df['fiscal_month'] = long_df['period'].apply(get_month)
        long_df['account_name'] = long_df[item_col]
        long_df = long_df[['fiscal_month', 'account_name', 'amount']]
        long_df = long_df[long_df['fiscal_month'] > 0]
        long_df = long_df[long_df['amount'].notna()]
        long_df['amount'] = pd.to_numeric(long_df['amount'], errors='coerce').fillna(0)
        
        return long_df
    except:
        return df


# シナリオ自動生成関数
def generate_scenario_data(base_data, scenario_type):
    """
    ベースデータからシナリオデータを自動生成
    
    Args:
        base_data: dict {項目名: 金額}
        scenario_type: 'optimistic' or 'pessimistic'
    
    Returns:
        dict: 調整後のデータ
    """
    result = base_data.copy()
    
    if scenario_type == 'optimistic':
        # 楽観シナリオ
        adjustments = {
            '売上高': 1.10,
            '売上': 1.10,
            '売上原価': 0.90,
            '原価': 0.90,
            '販売費及び一般管理費': 0.95,
            '販管費': 0.95,
            '一般管理費': 0.95
        }
    else:
        # 悲観シナリオ
        adjustments = {
            '売上高': 0.90,
            '売上': 0.90,
            '売上原価': 1.10,
            '原価': 1.10,
            '販売費及び一般管理費': 1.10,
            '販管費': 1.10,
            '一般管理費': 1.10
        }
    
    # 調整適用
    for item, value in result.items():
        for key, ratio in adjustments.items():
            if key in item:
                result[item] = value * ratio
                break
    
    return result


# サイドバー
st.sidebar.markdown("""
<div style='text-align: center; padding: 1rem 0;'>
    <h1 style='color: #1f77b4; margin: 0; font-size: 1.8rem;'>📊</h1>
    <h2 style='color: #2c3e50; margin: 0.5rem 0 0 0; font-size: 1.3rem;'>財務予測<br>シミュレーター</h2>
</div>
""", unsafe_allow_html=True)

st.sidebar.markdown("---")

# セッション状態の初期化
if 'authenticated' not in st.session_state:
    st.session_state.authenticated = True  # ログイン機能を一時的に無効化
if 'username' not in st.session_state:
    st.session_state.username = "ユーザー"

# ユーザー情報とログアウト
st.sidebar.markdown(f"**👤 {st.session_state.username}**")
if st.sidebar.button("ログアウト", type="secondary"):
    st.session_state.authenticated = False
    st.session_state.username = ""
    st.rerun()

st.sidebar.markdown("---")

# データベース接続状態の表示
if processor.use_postgres:
    st.sidebar.success("🌐 Supabase接続中")
else:
    st.sidebar.warning("💾 SQLite使用中")
    st.sidebar.caption("⚠️ データは一時的です")

st.sidebar.markdown("---")

# 会社選択
companies = get_companies_cached(processor)
if companies.empty:
    st.sidebar.info("🏢 会社を登録してください")
    st.sidebar.markdown("👉 システム設定から会社を追加")
    # 強制的にシステム設定ページに
    st.session_state.page = "システム設定"
    selected_comp_name = ""
    selected_comp_id = None
    
    # メニューを表示（システム設定のみ使用可能）
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 📋 メニュー")
    st.sidebar.markdown("⚙️ システム設定")
    
else:
    comp_names = companies['name'].tolist()
    
    # 前回の選択を保存
    prev_comp_id = st.session_state.get('selected_comp_id', None)
    
    selected_comp_name = st.sidebar.selectbox(
        "🏢 会社を選択",
        comp_names,
        key="comp_select"
    )
    selected_comp_id = int(companies[companies['name'] == selected_comp_name]['id'].iloc[0])
    
    # 会社が変更された場合、データをリフレッシュ
    if prev_comp_id != selected_comp_id:
        # session_stateをクリア（データ再読み込み用）
        for key in ['actuals_df', 'forecasts_df', 'imported_df', 'show_import_button']:
            if key in st.session_state:
                del st.session_state[key]
    
    st.session_state.selected_comp_id = selected_comp_id
    st.session_state.selected_comp_name = selected_comp_name

    # 期選択
    periods = get_company_periods_cached(selected_comp_id, processor)
    if periods.empty:
        st.sidebar.info("📅 会計期間を登録してください")
        st.sidebar.markdown("👉 システム設定から期を追加")
        selected_period_num = 0
        selected_period_id = None
    else:
        # 前回の選択を保存
        prev_period_id = st.session_state.get('selected_period_id', None)
        
        period_options = [
            f"第{row['period_num']}期 ({row['start_date']} 〜 {row['end_date']})"
            for _, row in periods.iterrows()
        ]
        selected_period_str = st.sidebar.selectbox(
            "📅 期を選択",
            period_options,
            key="period_select"
        )
        selected_period_num = int(selected_period_str.split('第')[1].split('期')[0])
        periods.columns = [c.lower() for c in periods.columns]
        
        period_match = periods[periods['period_num'] == selected_period_num]
        if not period_match.empty:
            if 'id' in period_match.columns:
                selected_period_id = int(period_match['id'].iloc[0])
            else:
                selected_period_id = int(period_match.iloc[0, 0])
            
            # 期が変更された場合、データをリフレッシュ
            if prev_period_id != selected_period_id:
                # session_stateをクリア（データ再読み込み用）
                for key in ['actuals_df', 'forecasts_df', 'imported_df', 'show_import_button']:
                    if key in st.session_state:
                        del st.session_state[key]
                
            st.session_state.selected_period_id = selected_period_id
            st.session_state.selected_period_num = selected_period_num
            st.session_state.start_date = period_match['start_date'].iloc[0]
            st.session_state.end_date = period_match['end_date'].iloc[0]
        else:
            st.error("選択された期が見つかりません")
            selected_period_id = None

    # 予測シナリオ
    st.sidebar.markdown("### 🎯 予測シナリオ")
    st.session_state.scenario = st.sidebar.radio(
        "シナリオを選択",
        ["現実", "楽観", "悲観"],
        horizontal=True,
        label_visibility="collapsed"
    )
    
    # シナリオ設定
    if 'scenario_rates' not in st.session_state:
        st.session_state.scenario_rates = {
            "現実": 0.0,
            "楽観": 0.1,
            "悲観": -0.1
        }
    
    # 表示設定
    st.sidebar.markdown("### ⚙️ 表示設定")
    st.session_state.display_mode = st.sidebar.radio(
        "表示モード",
        ["要約", "詳細"],
        horizontal=True
    )
    
    # 月次リスト取得
    if selected_period_id:
        months = get_fiscal_months_cached(selected_comp_id, selected_period_id, processor)
        
        # 実績締月の選択
        if 'current_month' not in st.session_state or st.session_state.current_month not in months:
            st.session_state.current_month = months[0]
            
        st.session_state.current_month = st.sidebar.selectbox(
            "実績締月を選択",
            months,
            index=months.index(st.session_state.current_month) if st.session_state.current_month in months else 0
        )

    # メニュー
    st.sidebar.markdown("---")
    
    # 階層型ナビゲーション（アイコンなし）
    st.sidebar.markdown("### ダッシュボード")
    if st.sidebar.button("📊 CFO意思決定支援", width="stretch", key="nav_cfo_dashboard"):
        st.session_state.page = "CFO意思決定支援ダッシュボード"
    if st.sidebar.button("着地予測（PL）", width="stretch", key="nav_dashboard"):
        st.session_state.page = "着地予測ダッシュボード"
    
    st.sidebar.markdown("---")
    st.sidebar.markdown("### データ入力")
    col1, col2 = st.sidebar.columns(2)
    with col1:
        if st.button("実績", width="stretch", key="nav_actual"):
            st.session_state.page = "実績データ入力"
    with col2:
        if st.button("予測", width="stretch", key="nav_forecast"):
            st.session_state.page = "予測データ入力"
    
    if st.sidebar.button("データ取込", width="stretch", key="nav_import"):
        st.session_state.page = "データインポート"
    
    if st.sidebar.button("シナリオ一括設定", width="stretch", key="nav_scenario"):
        st.session_state.page = "シナリオ一括設定"
    
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 財務諸表")
    if st.sidebar.button("損益計算書 (PL)", width="stretch", key="nav_pl"):
        st.session_state.page = "損益計算書 (PL)"
    if st.sidebar.button("貸借対照表 (BS)", width="stretch", key="nav_bs"):
        st.session_state.page = "貸借対照表 (BS)"
    if st.sidebar.button("CF計算書", width="stretch", key="nav_cf"):
        st.session_state.page = "キャッシュフロー計算書 (CF)"
    if st.sidebar.button("CF詳細分析", width="stretch", key="nav_cf_detail"):
        st.session_state.page = "CF詳細分析"
    
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 分析レポート")
    if st.sidebar.button("予実比較", width="stretch", key="nav_comparison"):
        st.session_state.page = "予測 VS 実績比較"
    if st.sidebar.button("シナリオ比較", width="stretch", key="nav_scenario_comp"):
        st.session_state.page = "シナリオ比較"
    if st.sidebar.button("期間比較", width="stretch", key="nav_period"):
        st.session_state.page = "期間比較分析"
    if st.sidebar.button("経営指標", width="stretch", key="nav_metrics"):
        st.session_state.page = "経営指標ダッシュボード"
    if st.sidebar.button("損益分岐点", width="stretch", key="nav_breakeven"):
        st.session_state.page = "損益分岐点分析"
    if st.sidebar.button("運転資本分析", width="stretch", key="nav_working_capital"):
        st.session_state.page = "運転資本分析"
    if st.sidebar.button("収益構造分析", width="stretch", key="nav_profitability"):
        st.session_state.page = "収益構造分析"
    
    # AI自動予測
    if ADVANCED_FORECAST_AVAILABLE:
        if st.sidebar.button("🔮 AI自動予測", width="stretch", key="nav_ai_forecast"):
            st.session_state.page = "AI自動予測"
    
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 設定")
    if st.sidebar.button("システム設定", width="stretch", key="nav_settings"):
        st.session_state.page = "システム設定"
    
    # ページ情報を保持（後方互換性のため）
    if 'page' not in st.session_state:
        st.session_state.page = "着地予測ダッシュボード"

# --------------------------------------------------------------------------------
# ヘルパー関数
# --------------------------------------------------------------------------------
def format_currency(val):
    """通貨フォーマット"""
    if pd.isna(val):
        return "¥0"
    return f"¥{safe_int(val):,}"

def format_percent(val):
    """パーセントフォーマット"""
    if pd.isna(val):
        return "0.0%"
    return f"{val:.1f}%"

# --------------------------------------------------------------------------------
# メインコンテンツ
# --------------------------------------------------------------------------------

# システム設定ページ（会社未登録時でも表示）
if st.session_state.page == "システム設定":
    st.title("システム設定")
    
    tab1, tab2, tab3 = st.tabs(["🏢 会社設定", "📅 会計期間設定", "🔍 データベース診断"])
    
    with tab1:
        st.subheader("会社情報の管理")
        
        # 新規会社登録
        with st.form("company_form"):
            new_company_name = st.text_input("新規会社名", placeholder="株式会社サンプル")
            if st.form_submit_button("➕ 会社を登録", type="primary"):
                if new_company_name:
                    success, msg = processor.register_company(new_company_name)
                    if success:
                        st.success(msg)
                        st.rerun()
                    else:
                        st.error(msg)
                else:
                    st.error("会社名を入力してください")
        
        st.markdown("---")
        
        # 登録済み会社一覧
        st.subheader("📋 登録済み会社")
        if not companies.empty:
            st.dataframe(companies, width=600)
        else:
            st.info("登録されている会社がありません")
            
    with tab2:
        st.subheader("会計期間の管理")
        
        if companies.empty:
            st.warning("先に会社を登録してください")
        else:
            comp_id_for_period = st.selectbox(
                "対象会社を選択",
                companies['id'].tolist(),
                format_func=lambda x: companies[companies['id'] == x]['name'].iloc[0]
            )
            
            with st.form("period_form"):
                col1, col2 = st.columns(2)
                with col1:
                    period_num = st.number_input("期数 (第n期)", min_value=1, value=1)
                with col2:
                    start_date = st.date_input("開始日")
                    end_date = st.date_input("終了日")
                
                if st.form_submit_button("➕ 期を追加", type="primary"):
                    if start_date and end_date:
                        if start_date < end_date:
                            success, msg = processor.register_fiscal_period(comp_id_for_period, period_num, str(start_date), str(end_date))
                            if success:
                                st.success(msg)
                                st.rerun()
                            else:
                                st.error(msg)
                        else:
                            st.error("❌ 終了日は開始日より後である必要があります")
                    else:
                        st.error("❌ すべてのフィールドを入力してください")
            
            st.markdown("---")
            
            # 登録済み期間一覧
            st.subheader("📋 登録済み会計期間")
            
            if 'selected_comp_id' in st.session_state and st.session_state.selected_comp_id:
                periods_list = processor.get_company_periods(st.session_state.selected_comp_id)
                if not periods_list.empty:
                    st.dataframe(periods_list, width=800)
                else:
                    st.info("登録されている会計期間がありません")
            else:
                st.info("会社を選択すると、その会社の期間が表示されます")
    
    with tab3:
        st.subheader("🔍 データベース診断")
        
        # 接続状態
        st.markdown("### 📡 接続状態")
        if processor.use_postgres:
            st.success("✅ **PostgreSQL (Supabase) 接続中**")
            st.markdown("""
            <div class="success-box">
                <strong>データは永続的に保存されます</strong><br>
                • アプリ再起動後もデータが残ります<br>
                • 複数デバイスから同じデータにアクセス可能<br>
                • データは安全にクラウドに保存されています
            </div>
            """, unsafe_allow_html=True)
            
            # Supabase設定情報
            if hasattr(st, 'secrets') and 'database' in st.secrets:
                st.markdown("### ⚙️ Supabase設定")
                config_info = {
                    "項目": ["ホスト", "データベース", "ユーザー", "ポート"],
                    "値": [
                        st.secrets['database']['host'],
                        st.secrets['database']['database'],
                        st.secrets['database']['user'],
                        str(st.secrets['database']['port'])
                    ]
                }
                st.table(pd.DataFrame(config_info))
        else:
            st.warning("⚠️ **SQLite ローカルデータベース使用中**")
            st.markdown("""
            <div class="warning-box">
                <strong>データは一時的です</strong><br>
                • Streamlit Cloudではアプリ再起動時にデータが消えます<br>
                • ローカル環境では問題なく動作します<br>
                • 永続化するにはSupabaseの設定が必要です
            </div>
            """, unsafe_allow_html=True)
        
        st.markdown("---")
        
        # データ統計
        st.markdown("### 📊 データ統計")
        
        companies_stat = processor.get_companies()
        total_companies = len(companies_stat)
        
        st.metric("登録会社数", f"{total_companies}社")
        
        if total_companies > 0 and 'selected_comp_id' in st.session_state and st.session_state.selected_comp_id:
            periods_stat = processor.get_company_periods(st.session_state.selected_comp_id)
            st.metric("会計期間数", f"{len(periods_stat)}期")
        
        # 接続テスト
        st.markdown("---")
        st.markdown("### 🧪 接続テスト")
        
        if st.button("🔄 データベース接続をテスト", type="primary"):
            with st.spinner("接続テスト中..."):
                try:
                    # 簡単なクエリで接続テスト
                    test_result = processor.get_companies()
                    st.success(f"✅ 接続成功！会社データを{len(test_result)}件取得しました")
                except Exception as e:
                    st.error(f"❌ 接続失敗: {str(e)}")

# データの読み込み（期が選択されている場合のみ）
if 'selected_period_id' in st.session_state and st.session_state.selected_period_id is not None:
        # キャッシュされたデータを使用
        with st.spinner('データを読み込んでいます...'):
            if 'actuals_df' not in st.session_state:
                st.session_state.actuals_df = load_actual_data_cached(st.session_state.selected_period_id, processor)
            if 'forecasts_df' not in st.session_state:
                st.session_state.forecasts_df = load_forecast_data_cached(st.session_state.selected_period_id, "現実", processor)
            if 'sub_accounts_df' not in st.session_state:
                st.session_state.sub_accounts_df = load_sub_accounts_cached(st.session_state.selected_period_id, st.session_state.scenario, processor)
            
        actuals_df = st.session_state.actuals_df.copy()
        forecasts_df = st.session_state.forecasts_df.copy()
        sub_accounts_df = st.session_state.sub_accounts_df.copy()
        
        # シナリオ調整（キャッシュ & ベクトル化）
        adjustment_key = (st.session_state.scenario, st.session_state.current_month)
        if st.session_state.scenario != "現実":
            if 'scenario_adjustment_cache' not in st.session_state or st.session_state.get('adjustment_key') != adjustment_key:
                rate = st.session_state.scenario_rates[st.session_state.scenario]
                split_idx = months.index(st.session_state.current_month) + 1 if st.session_state.current_month in months else 0
                forecast_months = months[split_idx:]
                # DataFrameに存在する月のみを使用
                available_forecast_months = [m for m in forecast_months if m in forecasts_df.columns]
                
                # ベクトル化: 条件に応じて一括調整
                if available_forecast_months:
                    # 売上高: +rate
                    forecasts_df.loc[forecasts_df['項目名'] == '売上高', available_forecast_months] *= (1 + rate)
                    
                    # 売上原価: -rate*0.5
                    forecasts_df.loc[forecasts_df['項目名'] == '売上原価', available_forecast_months] *= (1 - rate * 0.5)
                    
                    # 販管費: -rate*0.3 (一括)
                    ga_mask = forecasts_df['項目名'].isin(processor.ga_items)
                    forecasts_df.loc[ga_mask, available_forecast_months] *= (1 - rate * 0.3)
                
                st.session_state.scenario_adjustment_cache = forecasts_df.copy()
                st.session_state.adjustment_key = adjustment_key
            else:
                forecasts_df = st.session_state.scenario_adjustment_cache.copy()
        
        # 補助科目合計の反映（最適化）
        if not sub_accounts_df.empty:
            sub_cache_key = (st.session_state.selected_period_id, st.session_state.scenario)
            if 'sub_account_aggregation_cache' not in st.session_state or st.session_state.get('sub_cache_key') != sub_cache_key:
                # groupbyで集計（高速）
                aggregated = sub_accounts_df.groupby(['parent_item', 'month'])['amount'].sum().reset_index()
                
                # ピボットテーブルで一括更新（高速化）
                pivot = aggregated.pivot(index='parent_item', columns='month', values='amount')
                
                for parent in pivot.index:
                    mask = forecasts_df['項目名'] == parent
                    for month in pivot.columns:
                        if month in forecasts_df.columns:
                            forecasts_df.loc[mask, month] = pivot.loc[parent, month]
                
                st.session_state.sub_account_aggregation_cache = forecasts_df.copy()
                st.session_state.sub_cache_key = sub_cache_key
            else:
                forecasts_df = st.session_state.sub_account_aggregation_cache.copy()
        
        # PL計算（キャッシュ）
        if 'pl_df' not in st.session_state or st.session_state.get('pl_cache_key') != (st.session_state.selected_period_id, st.session_state.scenario, st.session_state.current_month):
            split_idx = months.index(st.session_state.current_month) + 1 if st.session_state.current_month in months else 0
            pl_df = processor.calculate_pl(
                actuals_df,
                forecasts_df,
                split_idx,
                months
            )
            st.session_state.pl_df = pl_df
            st.session_state.pl_cache_key = (st.session_state.selected_period_id, st.session_state.scenario, st.session_state.current_month)
        else:
            pl_df = st.session_state.pl_df
        
        # 表示モードでフィルタ
        if st.session_state.display_mode == "要約":
            pl_display = pl_df[pl_df['タイプ'] == '要約']
        else:
            pl_display = pl_df
        
        # --------------------------------------------------------------------------------
        # ページコンテンツ
        # --------------------------------------------------------------------------------
        
        if st.session_state.page == "着地予測ダッシュボード":
            st.title("財務予測シミュレーター")
            
            # ヘッダー（3列）
            col1, col2, col3 = st.columns([3, 2, 2])
            with col1:
                st.markdown(f"**{st.session_state.selected_comp_name}** | 第{st.session_state.selected_period_num}期")
            with col2:
                st.markdown(f"実績: {st.session_state.start_date} 〜 {st.session_state.current_month}")
            with col3:
                scenario_options = ["現実", "楽観", "悲観"]
                current_idx = scenario_options.index(st.session_state.scenario) if st.session_state.scenario in scenario_options else 0
                selected_scenario = st.selectbox(
                    "シナリオ切替", 
                    scenario_options,
                    index=current_idx,
                    key="dashboard_scenario_selector",
                    label_visibility="collapsed"
                )
                if selected_scenario != st.session_state.scenario:
                    st.session_state.scenario = selected_scenario
                    # キャッシュをクリア
                    for key in ['pl_df', 'forecast_data_cache', 'sub_account_aggregation_cache']:
                        if key in st.session_state:
                            del st.session_state[key]
                    st.rerun()
            
            st.markdown("---")
            
            # ダッシュボードヘッダー
            st.markdown(f"""
            <div class="dashboard-header fade-in">
                <h1 class="dashboard-title">📊 財務予測ダッシュボード</h1>
                <p class="dashboard-subtitle">{st.session_state.selected_comp_name} | 第{st.session_state.selected_period_num}期 | {st.session_state.start_date} 〜 {st.session_state.end_date} | シナリオ: {st.session_state.scenario}</p>
            </div>
            """, unsafe_allow_html=True)
            
            # 主要金額カード（3列）- マネージボード風
            col1, col2, col3 = st.columns(3)
            
            # 予測・実績・前年の計算
            sales_forecast = pl_display[pl_display['項目名'] == '売上高']['合計'].iloc[0] if not pl_display.empty else 0
            op_forecast = pl_display[pl_display['項目名'] == '営業損益金額']['合計'].iloc[0] if not pl_display.empty else 0
            ord_forecast = pl_display[pl_display['項目名'] == '経常損益金額']['合計'].iloc[0] if not pl_display.empty else 0
            
            # 実績（現在月まで）
            actual_months = [m for m in months if m <= st.session_state.current_month]
            sales_actual = 0
            op_actual = 0
            ord_actual = 0
            
            if not actuals_df.empty and actual_months:
                sales_row = actuals_df[actuals_df['項目名'] == '売上高']
                op_row = actuals_df[actuals_df['項目名'] == '営業損益金額']
                ord_row = actuals_df[actuals_df['項目名'] == '経常損益金額']
                
                if not sales_row.empty:
                    for m in actual_months:
                        if m in sales_row.columns:
                            val = sales_row[m].iloc[0]
                            sales_actual += float(val) if pd.notna(val) else 0
                
                if not op_row.empty:
                    for m in actual_months:
                        if m in op_row.columns:
                            val = op_row[m].iloc[0]
                            op_actual += float(val) if pd.notna(val) else 0
                
                if not ord_row.empty:
                    for m in actual_months:
                        if m in ord_row.columns:
                            val = ord_row[m].iloc[0]
                            ord_actual += float(val) if pd.notna(val) else 0
            
            # 達成率計算
            sales_achievement = (sales_actual / sales_forecast * 100) if sales_forecast != 0 else 0
            op_achievement = (op_actual / op_forecast * 100) if op_forecast != 0 else 0
            ord_achievement = (ord_actual / ord_forecast * 100) if ord_forecast != 0 else 0
            
            # KPIカード表示（マネージボード風）
            with col1:
                trend_icon = "▲" if sales_achievement >= 0 else "▼"
                st.markdown(f"""
                <div class="kpi-card fade-in" style="background: linear-gradient(135deg, #3B82F6 0%, #60A5FA 100%);">
                    <div class="kpi-card-title">売上高</div>
                    <div class="kpi-card-value">¥{safe_int(sales_forecast):,}</div>
                    <div class="kpi-card-subtitle">通期予測</div>
                    <div class="kpi-card-trend">
                        <span>{trend_icon}</span>
                        <span>実績: ¥{safe_int(sales_actual):,}</span>
                        <span style="margin-left: 8px;">({sales_achievement:.1f}%)</span>
                    </div>
                    <div class="kpi-card-decoration"></div>
                </div>
                """, unsafe_allow_html=True)
            
            with col2:
                trend_icon = "▲" if op_achievement >= 0 else "▼"
                trend_color = "#10B981" if op_forecast >= 0 else "#EF4444"
                st.markdown(f"""
                <div class="kpi-card fade-in" style="background: linear-gradient(135deg, #10B981 0%, #34D399 100%);">
                    <div class="kpi-card-title">営業利益</div>
                    <div class="kpi-card-value">¥{safe_int(op_forecast):,}</div>
                    <div class="kpi-card-subtitle">通期予測</div>
                    <div class="kpi-card-trend">
                        <span>{trend_icon}</span>
                        <span>実績: ¥{safe_int(op_actual):,}</span>
                        <span style="margin-left: 8px;">({op_achievement:.1f}%)</span>
                    </div>
                    <div class="kpi-card-decoration"></div>
                </div>
                """, unsafe_allow_html=True)
            
            with col3:
                trend_icon = "▲" if ord_achievement >= 0 else "▼"
                trend_color = "#10B981" if ord_forecast >= 0 else "#EF4444"
                st.markdown(f"""
                <div class="kpi-card fade-in" style="background: linear-gradient(135deg, #F59E0B 0%, #FBBF24 100%);">
                    <div class="kpi-card-title">経常利益</div>
                    <div class="kpi-card-value">¥{safe_int(ord_forecast):,}</div>
                    <div class="kpi-card-subtitle">通期予測</div>
                    <div class="kpi-card-trend">
                        <span>{trend_icon}</span>
                        <span>実績: ¥{safe_int(ord_actual):,}</span>
                        <span style="margin-left: 8px;">({ord_achievement:.1f}%)</span>
                    </div>
                    <div class="kpi-card-decoration"></div>
                </div>
                """, unsafe_allow_html=True)
            
            # スペース
            st.markdown("<div style='height: 24px;'></div>", unsafe_allow_html=True)
            
            # 月次推移グラフ（Manageboard風）
            st.markdown('<div class="section-card fade-in">', unsafe_allow_html=True)
            st.markdown('<h3 class="section-title">📈 月次推移</h3>', unsafe_allow_html=True)
            
            # 実績と予測を分ける
            actual_months_list = [m for m in months if m <= st.session_state.current_month]
            forecast_months_list = [m for m in months if m > st.session_state.current_month]
            
            # データ取得
            sales_row = pl_df[pl_df['項目名'] == '売上高']
            op_row = pl_df[pl_df['項目名'] == '営業損益金額']
            
            if not sales_row.empty:
                # 実績データ
                sales_actual_data = []
                for m in actual_months_list:
                    if m in sales_row.columns:
                        val = sales_row[m].iloc[0]
                        sales_actual_data.append(float(val) if pd.notna(val) else 0)
                    else:
                        sales_actual_data.append(0)
                
                # 予測データ
                sales_forecast_data = []
                for m in forecast_months_list:
                    if m in sales_row.columns:
                        val = sales_row[m].iloc[0]
                        sales_forecast_data.append(float(val) if pd.notna(val) else 0)
                    else:
                        sales_forecast_data.append(0)
                
                # 営業利益データ
                op_data = []
                for m in months:
                    if m in op_row.columns:
                        val = op_row[m].iloc[0]
                        op_data.append(float(val) if pd.notna(val) else 0)
                    else:
                        op_data.append(0)
                
                # グラフ作成（Manageboard風カラー）
                fig = make_subplots(specs=[[{"secondary_y": True}]])
                
                # 実績（濃い鮮やかなブルー）
                fig.add_trace(
                    go.Bar(
                        x=actual_months_list,
                        y=sales_actual_data,
                        name="売上高（実績）",
                        marker_color='#2563eb',
                        opacity=1.0,
                        hovertemplate='%{x}<br>¥%{y:,.0f}<extra></extra>'
                    ),
                    secondary_y=False
                )
                
                # 予測（明るいシアン + グラデーション効果）
                fig.add_trace(
                    go.Bar(
                        x=forecast_months_list,
                        y=sales_forecast_data,
                        name="売上高（予測）",
                        marker=dict(
                            color='#60a5fa',
                            line=dict(color='#3b82f6', width=0)
                        ),
                        opacity=0.7,
                        hovertemplate='%{x}<br>¥%{y:,.0f}<extra></extra>'
                    ),
                    secondary_y=False
                )
                
                # 営業利益（太い線 + 丸いマーカー）
                fig.add_trace(
                    go.Scatter(
                        x=months,
                        y=op_data,
                        name="営業利益",
                        line=dict(color='#10B981', width=4),
                        mode='lines+markers',
                        marker=dict(
                            size=10,
                            color='#10B981',
                            line=dict(color='white', width=2)
                        ),
                        hovertemplate='%{x}<br>¥%{y:,.0f}<extra></extra>'
                    ),
                    secondary_y=True
                )
                
                # レイアウト - マネージボード風
                fig.update_layout(
                    template='plotly_white',
                    paper_bgcolor='rgba(0,0,0,0)',
                    plot_bgcolor='rgba(0,0,0,0)',
                    font=dict(
                        family="Inter, -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif",
                        size=13,
                        color='#1F2937'
                    ),
                    xaxis=dict(
                        title="",
                        showgrid=False,
                        showline=True,
                        linewidth=1,
                        linecolor='#E5E7EB',
                        tickfont=dict(color='#6B7280', size=12)
                    ),
                    yaxis=dict(
                        title="売上高",
                        title_font=dict(size=14, color='#374151', weight=600),
                        showgrid=True,
                        gridwidth=1,
                        gridcolor='#F3F4F6',
                        showline=False,
                        tickfont=dict(color='#6B7280', size=12),
                        tickformat=',.0f'
                    ),
                    yaxis2=dict(
                        title="営業利益",
                        title_font=dict(size=14, color='#374151', weight=600),
                        overlaying='y',
                        side='right',
                        showgrid=False,
                        showline=False,
                        tickfont=dict(color='#6B7280', size=12),
                        tickformat=',.0f'
                    ),
                    legend=dict(
                        bgcolor='rgba(255, 255, 255, 0.8)',
                        bordercolor='#E5E7EB',
                        borderwidth=1,
                        font=dict(size=12, color='#1F2937'),
                        orientation="h",
                        yanchor="bottom",
                        y=1.02,
                        xanchor="right",
                        x=1
                    ),
                    hovermode='x unified',
                    height=450,
                    barmode='group',
                    margin=dict(l=20, r=20, t=40, b=20)
                )
                
                st.plotly_chart(fig, width="stretch")
            
            st.markdown('</div>', unsafe_allow_html=True)  # section-card終了
            
            # スペース
            st.markdown("<div style='height: 24px;'></div>", unsafe_allow_html=True)
            
            # ワンクリック着地予測ボタン
            col1, col2 = st.columns([1, 3])
            with col1:
                show_forecast = st.button("▶ 着地予測を表示", type="primary", width="stretch")
            
            if show_forecast or st.session_state.get('show_forecast_detail', False):
                st.session_state.show_forecast_detail = True
                
                with st.container():
                    st.markdown("### 📊 通期着地予測")
                    
                    # 達成率の計算（仮の目標値）
                    # TODO: 実際の目標値をデータベースから取得
                    target_sales = sales_forecast * 0.98  # 仮の目標（予測の98%）
                    achievement_rate = (sales_forecast / target_sales * 100) if target_sales != 0 else 0
                    
                    # 予測 vs 計画
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        plan_diff = 2.0  # 仮の計画比（％）
                        st.metric(
                            "売上高（通期）",
                            f"¥{safe_int(sales_forecast):,}",
                            delta=f"{plan_diff:+.1f}% vs 計画"
                        )
                    
                    with col2:
                        plan_diff_op = 5.0  # 仮の計画比
                        st.metric(
                            "営業利益（通期）",
                            f"¥{safe_int(op_forecast):,}",
                            delta=f"{plan_diff_op:+.1f}% vs 計画"
                        )
                    
                    with col3:
                        plan_diff_ord = 3.0  # 仮の計画比
                        st.metric(
                            "経常利益（通期）",
                            f"¥{safe_int(ord_forecast):,}",
                            delta=f"{plan_diff_ord:+.1f}% vs 計画"
                        )
                    
                    # 達成率
                    st.markdown("---")
                    st.markdown("#### 🎯 目標達成状況")
                    
                    col1, col2 = st.columns([2, 1])
                    with col1:
                        st.progress(min(achievement_rate / 100, 1.0))
                    with col2:
                        achievement_color = "🟢" if achievement_rate >= 100 else "🟡" if achievement_rate >= 90 else "🔴"
                        st.markdown(f"### {achievement_color} {achievement_rate:.1f}%")
                    
                    st.caption(f"目標売上: ¥{safe_int(target_sales):,}")
            
            st.markdown("---")
            
            # 主要指標（ドリルダウン対応）
            st.markdown("### 主要指標（クリックで詳細表示）")
            
            col1, col2, col3 = st.columns(3)
            
            gp_total = pl_display[pl_display['項目名'] == '売上総損益金額']['合計'].iloc[0] if not pl_display.empty else 0
            cogs_total = pl_display[pl_display['項目名'] == '売上原価']['合計'].iloc[0] if not pl_display.empty else 0
            
            gp_rate = ((sales_forecast - cogs_total) / sales_forecast * 100) if sales_forecast != 0 else 0
            op_rate = (op_forecast / sales_forecast * 100) if sales_forecast != 0 else 0
            ord_rate = (ord_forecast / sales_forecast * 100) if sales_forecast != 0 else 0
            
            with col1:
                if st.button(f"粗利率: {gp_rate:.1f}%", width="stretch", key="drill_gp"):
                    st.session_state.drill_item = "粗利率"
            with col2:
                if st.button(f"営業利益率: {op_rate:.1f}%", width="stretch", key="drill_op"):
                    st.session_state.drill_item = "営業利益率"
            with col3:
                if st.button(f"経常利益率: {ord_rate:.1f}%", width="stretch", key="drill_ord"):
                    st.session_state.drill_item = "経常利益率"
            
            # ドリルダウン表示
            if 'drill_item' in st.session_state and st.session_state.drill_item:
                with st.expander(f"📊 {st.session_state.drill_item}の詳細", expanded=True):
                    if st.session_state.drill_item == "粗利率":
                        st.markdown(f"""
                        **計算式:**  
                        粗利率 = (売上高 - 売上原価) ÷ 売上高  
                        = (¥{safe_int(sales_forecast):,} - ¥{safe_int(cogs_total):,}) ÷ ¥{safe_int(sales_forecast):,}  
                        = **{gp_rate:.1f}%**
                        
                        **内訳:**
                        - 売上高: ¥{safe_int(sales_forecast):,}
                        - 売上原価: ¥{safe_int(cogs_total):,}
                        - 売上総利益: ¥{safe_int(gp_total):,}
                        """)
                        
                        # 補助科目があれば表示
                        if not sub_accounts_df.empty:
                            sales_subs = sub_accounts_df[sub_accounts_df['parent_item'] == '売上高']
                            if not sales_subs.empty:
                                st.markdown("**売上高の内訳:**")
                                sub_summary = sales_subs.groupby('sub_account_name')['amount'].sum().reset_index()
                                sub_summary['構成比'] = (sub_summary['amount'] / sub_summary['amount'].sum() * 100).round(1)
                                
                                for _, row in sub_summary.iterrows():
                                    st.markdown(f"- {row['sub_account_name']}: ¥{safe_int(row['amount']):,} ({row['構成比']}%)")
                    
                    elif st.session_state.drill_item == "営業利益率":
                        ga_total = pl_display[pl_display['項目名'] == '販売管理費計']['合計'].iloc[0] if not pl_display.empty else 0
                        
                        st.markdown(f"""
                        **計算式:**  
                        営業利益率 = 営業利益 ÷ 売上高  
                        = ¥{safe_int(op_forecast):,} ÷ ¥{safe_int(sales_forecast):,}  
                        = **{op_rate:.1f}%**
                        
                        **内訳:**
                        - 売上総利益: ¥{safe_int(gp_total):,} ({gp_rate:.1f}%)
                        - 販売管理費: ¥{safe_int(ga_total):,} ({ga_total/sales_forecast*100:.1f}%)
                        - 営業利益: ¥{safe_int(op_forecast):,} ({op_rate:.1f}%)
                        """)
                    
                    elif st.session_state.drill_item == "経常利益率":
                        non_op_inc = pl_display[pl_display['項目名'] == '営業外収益合計']['合計'].iloc[0] if not pl_display.empty else 0
                        non_op_exp = pl_display[pl_display['項目名'] == '営業外費用合計']['合計'].iloc[0] if not pl_display.empty else 0
                        
                        st.markdown(f"""
                        **計算式:**  
                        経常利益率 = 経常利益 ÷ 売上高  
                        = ¥{safe_int(ord_forecast):,} ÷ ¥{safe_int(sales_forecast):,}  
                        = **{ord_rate:.1f}%**
                        
                        **内訳:**
                        - 営業利益: ¥{safe_int(op_forecast):,}
                        - 営業外収益: ¥{safe_int(non_op_inc):,}
                        - 営業外費用: ¥{safe_int(non_op_exp):,}
                        - 経常利益: ¥{safe_int(ord_forecast):,}
                        """)
                    
                    if st.button("閉じる", key="close_drill"):
                        st.session_state.drill_item = None
                        st.rerun()

        elif st.session_state.page == "CFO意思決定支援ダッシュボード":
            # ダッシュボードヘッダー
            st.markdown(f"""
            <div class="dashboard-header fade-in">
                <h1 class="dashboard-title">💰 CFO意思決定支援ダッシュボード</h1>
                <p class="dashboard-subtitle">{st.session_state.selected_comp_name} | 第{st.session_state.selected_period_num}期 | {st.session_state.start_date} 〜 {st.session_state.end_date}</p>
            </div>
            """, unsafe_allow_html=True)
            
            # アラート表示エリア
            st.markdown("<div style='height: 16px;'></div>", unsafe_allow_html=True)
            
                                                            # 実データからKPIを計算
            try:
                # デバッグ情報
                with st.expander("🔍 CFOデバッグ情報", expanded=False):
                    st.write("selected_period_id:", selected_period_id)
                    st.write("current_month:", st.session_state.current_month)
                
                # データ読み込み
                actuals_df = load_actual_data_cached(selected_period_id, processor)
                
                st.write("actuals_df loaded:", actuals_df is not None)
                if actuals_df is not None:
                    st.write("Shape:", actuals_df.shape)
                    st.write("Columns:", actuals_df.columns.tolist())
                
                if actuals_df is None or actuals_df.empty:
                    # データなし
                    use_real_data = False
                    operating_cf = 31500000
                    cash_balance = 376343476
                    cash_runway = 8.5
                    forecast_3months = 380000000
                    cf_growth = 8.3
                    alerts = [{'level': 'info', 'message': '💡 実績データを入力してください'}]
                
                else:
                    # 横持ち変換
                    actuals_df = convert_wide_to_long(actuals_df)
                    
                    st.write("After conversion:", actuals_df.shape)
                    st.dataframe(actuals_df.head())
                    
                    # カラム名正規化
                    actuals_df.columns = actuals_df.columns.str.lower()
                    
                    # カラム検出
                    month_col = account_col = amount_col = None
                    for col in ['fiscal_month', 'month', '月']:
                        if col in actuals_df.columns:
                            month_col = col
                            break
                    for col in ['account_name', 'account', '科目', '項目名']:
                        if col in actuals_df.columns:
                            account_col = col
                            break
                    for col in ['amount', '金額', 'value']:
                        if col in actuals_df.columns:
                            amount_col = col
                            break
                    
                    st.write(f"Detected: month={month_col}, account={account_col}, amount={amount_col}")
                    
                    if not all([month_col, account_col, amount_col]):
                        # カラムなし
                        use_real_data = False
                        operating_cf = 31500000
                        cash_balance = 376343476
                        cash_runway = 8.5
                        forecast_3months = 380000000
                        cf_growth = 8.3
                        alerts = [{'level': 'warning', 'message': f'カラム検出失敗'}]
                    else:
                        # 計算実行
                        # current_monthから月番号を抽出
                        current_month_str = str(st.session_state.current_month)
                        if '-' in current_month_str:
                            current_month = int(current_month_str.split('-')[-1])
                        elif '/' in current_month_str:
                            current_month = int(current_month_str.split('/')[-1])
                        else:
                            current_month = int(current_month_str)
                        
                        st.write(f"Extracted month: {current_month}")
                        
                        # 営業CF
                        operating_profit = 0
                        current_data = actuals_df[actuals_df[month_col] == current_month]
                        if not current_data.empty:
                            profit_rows = current_data[
                                current_data[account_col].astype(str).str.contains('営業', na=False)
                            ]
                            if not profit_rows.empty:
                                operating_profit = float(profit_rows[amount_col].iloc[0])
                        operating_cf = operating_profit * 0.8
                        
                        # 現金残高
                        cash_rows = actuals_df[
                            actuals_df[account_col].astype(str).str.contains('現金|預金', na=False)
                        ]
                        cash_balance = float(cash_rows[amount_col].sum()) if not cash_rows.empty else 100000000
                        
                        # 固定費
                        sg_rows = actuals_df[
                            actuals_df[account_col].astype(str).str.contains('販売費|販管費|一般管理費', na=False)
                        ]
                        if not sg_rows.empty:
                            total_sg = float(sg_rows[amount_col].sum())
                            months_count = max(actuals_df[month_col].nunique(), 1)
                            fixed_cost_monthly = total_sg / months_count
                        else:
                            fixed_cost_monthly = 40000000
                        
                        # KPI
                        cash_runway = cash_balance / fixed_cost_monthly if fixed_cost_monthly > 0 else 99
                        forecast_3months = cash_balance + (operating_cf * 3)
                        
                        # 前月比
                        prev_month = current_month - 1 if current_month > 1 else 12
                        prev_data = actuals_df[actuals_df[month_col] == prev_month]
                        cf_growth = 0
                        if not prev_data.empty:
                            prev_profit_rows = prev_data[
                                prev_data[account_col].astype(str).str.contains('営業', na=False)
                            ]
                            if not prev_profit_rows.empty:
                                prev_profit = float(prev_profit_rows[amount_col].iloc[0])
                                prev_cf = prev_profit * 0.8
                                cf_growth = ((operating_cf - prev_cf) / abs(prev_cf) * 100) if prev_cf != 0 else 0
                        
                        # アラート
                        alerts = []
                        if cash_runway < 3:
                            alerts.append({'level': 'critical', 'message': f'⚠️ 資金耐久: {cash_runway:.1f}ヶ月'})
                        elif cash_runway < 6:
                            alerts.append({'level': 'warning', 'message': f'資金耐久: {cash_runway:.1f}ヶ月'})
                        if operating_cf < 0:
                            alerts.append({'level': 'critical', 'message': '営業CFマイナス'})
                        
                        use_real_data = True
            
            except Exception as e:
                import sys
                import traceback
                sys.stderr.write(f"CFO error: {str(e)}\n")
                sys.stderr.write(traceback.format_exc())
                
                # デバッグ表示
                with st.expander("❌ エラー詳細", expanded=True):
                    st.error(f"エラー: {str(e)}")
                    st.code(traceback.format_exc())
                
                use_real_data = False
                operating_cf = 31500000
                cash_balance = 376343476
                cash_runway = 8.5
                forecast_3months = 380000000
                cf_growth = 8.3
                alerts = [{'level': 'warning', 'message': '⚠️ データ処理エラー'}]
            
            # アラート表示
            for alert in alerts:
                if alert['level'] == 'critical':
                    st.error(f"🚨 **資金危険**: {alert['message']}")
                elif alert['level'] == 'warning':
                    st.warning(f"⚠️ **資金注意**: {alert['message']}")
                else:
                    st.info(f"💡 {alert['message']}")
            
            st.markdown("<div style='height: 24px;'></div>", unsafe_allow_html=True)
            
            # KPIカード（3列）
            col1, col2, col3 = st.columns(3)
            
            # KPIカード表示（実データまたはサンプル）
            with col1:
                trend_icon = "▲" if cf_growth >= 0 else "▼"
                st.markdown(f"""
                <div class="kpi-card fade-in" style="background: linear-gradient(135deg, #3B82F6 0%, #60A5FA 100%);">
                    <div class="kpi-card-title">税引後キャッシュフロー</div>
                    <div class="kpi-card-value">¥{safe_int(operating_cf):,}</div>
                    <div class="kpi-card-subtitle">当月実績</div>
                    <div class="kpi-card-trend">
                        <span>{trend_icon}</span>
                        <span>前月比: {cf_growth:+.1f}%</span>
                    </div>
                    <div class="kpi-card-decoration"></div>
                </div>
                """, unsafe_allow_html=True)
            
            with col2:
                forecast_change = ((forecast_3months - cash_balance) / cash_balance * 100) if cash_balance != 0 else 0
                trend_icon = "▲" if forecast_change >= 0 else "▼"
                st.markdown(f"""
                <div class="kpi-card fade-in" style="background: linear-gradient(135deg, #10B981 0%, #34D399 100%);">
                    <div class="kpi-card-title">3ヶ月後現金残高（予測）</div>
                    <div class="kpi-card-value">¥{safe_int(forecast_3months):,}</div>
                    <div class="kpi-card-subtitle">標準シナリオ</div>
                    <div class="kpi-card-trend">
                        <span>{trend_icon}</span>
                        <span>現在比: {forecast_change:+.1f}%</span>
                    </div>
                    <div class="kpi-card-decoration"></div>
                </div>
                """, unsafe_allow_html=True)
            
            with col3:
                runway_status = "⚠️ 注意" if 6 <= cash_runway < 12 else ("🚨 危険" if cash_runway < 6 else "✅ 安全")
                st.markdown(f"""
                <div class="kpi-card fade-in" style="background: linear-gradient(135deg, #F59E0B 0%, #FBBF24 100%);">
                    <div class="kpi-card-title">資金耐久月数</div>
                    <div class="kpi-card-value">{cash_runway:.1f}ヶ月</div>
                    <div class="kpi-card-subtitle">現金残高 ÷ 月間固定費</div>
                    <div class="kpi-card-trend">
                        <span style="color: #F59E0B;">{runway_status}</span>
                    </div>
                    <div class="kpi-card-decoration"></div>
                </div>
                """, unsafe_allow_html=True)
            
            st.markdown("<div style='height: 32px;'></div>", unsafe_allow_html=True)
            
            # キャッシュフロー推移と予測グラフ
            st.markdown('<div class="section-card fade-in">', unsafe_allow_html=True)
            st.markdown('<h3 class="section-title">📈 キャッシュフロー推移と予測</h3>', unsafe_allow_html=True)
            
            # 実データから月次CFを作成
            try:
                if actuals_df is not None and not actuals_df.empty and month_col and account_col and amount_col:
                    # 月ごとの営業利益を取得
                    monthly_cf = []
                    month_labels = []
                    
                    for m in range(1, 13):
                        month_data = actuals_df[actuals_df[month_col] == m]
                        if not month_data.empty:
                            profit_rows = month_data[
                                month_data[account_col].astype(str).str.contains('営業', na=False)
                            ]
                            if not profit_rows.empty:
                                profit = float(profit_rows[amount_col].iloc[0])
                                monthly_cf.append(profit * 0.8)  # CF換算
                                month_labels.append(f"{m}月")
                    
                    # データがある場合は実データを使用
                    if monthly_cf:
                        # 実績データ（現在月まで）
                        current_month_num = int(st.session_state.current_month.split('-')[-1]) if '-' in str(st.session_state.current_month) else int(st.session_state.current_month)
                        
                        chart_months = month_labels[:12]
                        chart_actual = monthly_cf + [None] * (12 - len(monthly_cf))
                        
                        # 予測（現在月以降）
                        last_cf = monthly_cf[-1] if monthly_cf else operating_cf
                        forecast_start_idx = len(monthly_cf)
                        
                        chart_forecast_std = [None] * forecast_start_idx
                        chart_forecast_opt = [None] * forecast_start_idx
                        chart_forecast_pes = [None] * forecast_start_idx
                        
                        for i in range(forecast_start_idx, 12):
                            growth_factor = (i - forecast_start_idx + 1) * 0.02
                            chart_forecast_std.append(last_cf * (1 + growth_factor))
                            chart_forecast_opt.append(last_cf * 1.15 * (1 + growth_factor))
                            chart_forecast_pes.append(last_cf * 0.85 * (1 - growth_factor * 0.5))
                    else:
                        # データなし - サンプル使用
                        raise ValueError("No monthly CF data")
                else:
                    raise ValueError("No actuals data")
            except:
                # サンプルデータを使用
                chart_months = ['3月', '4月', '5月', '6月', '7月', '8月', '9月', '10月', '11月', '12月', '1月', '2月']
                chart_actual = [35000000, 28000000, 32000000, 31000000, 29000000, operating_cf, None, None, None, None, None, None]
                chart_forecast_std = [None, None, None, None, None, operating_cf, operating_cf*1.02, operating_cf*1.04, operating_cf*1.06, operating_cf*1.08, operating_cf*1.10, operating_cf*1.12]
                chart_forecast_opt = [None, None, None, None, None, operating_cf, operating_cf*1.15, operating_cf*1.20, operating_cf*1.25, operating_cf*1.30, operating_cf*1.35, operating_cf*1.40]
                chart_forecast_pes = [None, None, None, None, None, operating_cf, operating_cf*0.95, operating_cf*0.90, operating_cf*0.85, operating_cf*0.80, operating_cf*0.75, operating_cf*0.70]
            
            fig = go.Figure()
            
            # 実績（太い線）
            fig.add_trace(go.Scatter(
                x=chart_months,
                y=chart_actual,
                name='実績',
                mode='lines+markers',
                line=dict(color='#3B82F6', width=4),
                marker=dict(size=10, color='#3B82F6', line=dict(color='white', width=2)),
                hovertemplate='%{x}<br>¥%{y:,.0f}<extra></extra>'
            ))
            
            # 予測-標準（破線）
            fig.add_trace(go.Scatter(
                x=chart_months,
                y=chart_forecast_std,
                name='予測（標準）',
                mode='lines+markers',
                line=dict(color='#10B981', width=3, dash='dash'),
                marker=dict(size=8, color='#10B981'),
                hovertemplate='%{x}<br>¥%{y:,.0f}<extra></extra>'
            ))
            
            # 予測-楽観（細線）
            fig.add_trace(go.Scatter(
                x=chart_months,
                y=chart_forecast_opt,
                name='予測（楽観）',
                mode='lines',
                line=dict(color='#34D399', width=2, dash='dot'),
                hovertemplate='%{x}<br>¥%{y:,.0f}<extra></extra>'
            ))
            
            # 予測-悲観（細線）
            fig.add_trace(go.Scatter(
                x=chart_months,
                y=chart_forecast_pes,
                name='予測（悲観）',
                mode='lines',
                line=dict(color='#F59E0B', width=2, dash='dot'),
                hovertemplate='%{x}<br>¥%{y:,.0f}<extra></extra>'
            ))
            
            fig.update_layout(
                template='plotly_white',
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='rgba(0,0,0,0)',
                font=dict(family="Inter, sans-serif", size=13, color='#1F2937'),
                xaxis=dict(
                    title="",
                    showgrid=False,
                    showline=True,
                    linewidth=1,
                    linecolor='#E5E7EB',
                    tickfont=dict(color='#6B7280', size=12)
                ),
                yaxis=dict(
                    title="営業キャッシュフロー",
                    title_font=dict(size=14, color='#374151'),
                    showgrid=True,
                    gridwidth=1,
                    gridcolor='#F3F4F6',
                    showline=False,
                    tickfont=dict(color='#6B7280', size=12),
                    tickformat=',.0f'
                ),
                legend=dict(
                    bgcolor='rgba(255, 255, 255, 0.8)',
                    bordercolor='#E5E7EB',
                    borderwidth=1,
                    font=dict(size=12, color='#1F2937'),
                    orientation="h",
                    yanchor="bottom",
                    y=1.02,
                    xanchor="right",
                    x=1
                ),
                hovermode='x unified',
                height=400,
                margin=dict(l=20, r=20, t=40, b=20)
            )
            
            st.plotly_chart(fig, width="stretch")
            st.markdown('</div>', unsafe_allow_html=True)
            
            st.markdown("<div style='height: 24px;'></div>", unsafe_allow_html=True)
            
            # 2カラムセクション
            col1, col2 = st.columns(2)
            
            # CF内訳データを準備（実データから取得）
            cf_operating = operating_cf
            
            # デバッグ情報
            with st.expander("🔍 CF内訳デバッグ", expanded=False):
                st.write("**データ確認:**")
                st.write("actuals_df exists:", actuals_df is not None)
                if actuals_df is not None:
                    st.write("Shape:", actuals_df.shape)
                    st.write("Current month:", st.session_state.current_month)
            
            # 投資CF・財務CFを実データから取得
            try:
                if actuals_df is not None and not actuals_df.empty:
                    current_month_num = int(str(st.session_state.current_month).split('-')[-1]) if '-' in str(st.session_state.current_month) else int(st.session_state.current_month)
                    current_data = actuals_df[actuals_df[month_col] == current_month_num]
                    
                    st.write(f"Current month num: {current_month_num}")
                    st.write(f"Current data rows: {len(current_data)}")
                    st.write("Available accounts:", current_data[account_col].unique().tolist() if not current_data.empty else [])
                    
                    # 投資CF（設備投資、固定資産取得など）
                    investing_rows = current_data[
                        current_data[account_col].astype(str).str.contains('設備投資|固定資産|投資', na=False)
                    ]
                    cf_investing = -abs(float(investing_rows[amount_col].sum())) if not investing_rows.empty else 0
                    
                    st.write(f"Investing rows found: {len(investing_rows)}")
                    st.write(f"cf_investing: {cf_investing}")
                    
                    # 財務CF（借入、返済など）
                    financing_rows = current_data[
                        current_data[account_col].astype(str).str.contains('借入|返済|増資|配当', na=False)
                    ]
                    cf_financing = float(financing_rows[amount_col].sum()) if not financing_rows.empty else 0
                    
                    st.write(f"Financing rows found: {len(financing_rows)}")
                    st.write(f"cf_financing: {cf_financing}")
                    
                    # 現金増減
                    cf_net_change = cf_operating + cf_investing + cf_financing
                    
                    st.write(f"**計算結果:**")
                    st.write(f"- 営業CF: ¥{cf_operating:,.0f}")
                    st.write(f"- 投資CF: ¥{cf_investing:,.0f}")
                    st.write(f"- 財務CF: ¥{cf_financing:,.0f}")
                    st.write(f"- 現金増減: ¥{cf_net_change:,.0f}")
                else:
                    # データなし
                    cf_investing = 0
                    cf_financing = 0
                    cf_net_change = cf_operating
                    st.write("データなし - ゼロ表示")
            except Exception as e:
                # エラー時
                import traceback
                st.error(f"エラー: {e}")
                st.code(traceback.format_exc())
                cf_investing = 0
                cf_financing = 0
                cf_net_change = cf_operating
            
            with col1:
                st.markdown('<div class="section-card fade-in">', unsafe_allow_html=True)
                st.markdown('<h3 class="section-title">💰 キャッシュフロー内訳</h3>', unsafe_allow_html=True)
                
                operating_trend = "▲ +5.2%" if cf_operating > 0 else "▼"
                operating_class = "metric-up" if cf_operating > 0 else "metric-down"
                
                # CF内訳
                st.markdown(f"""
                <div class="metric-row">
                    <div class="metric-name">営業CF</div>
                    <div>
                        <span class="metric-value">¥{safe_int(cf_operating):,}</span>
                        <span class="metric-change {operating_class}">{operating_trend}</span>
                    </div>
                </div>
                <div class="metric-row">
                    <div class="metric-name">投資CF</div>
                    <div>
                        <span class="metric-value">¥{safe_int(cf_investing):,}</span>
                        <span class="metric-change metric-down">▼ 設備投資</span>
                    </div>
                </div>
                <div class="metric-row">
                    <div class="metric-name">財務CF</div>
                    <div>
                        <span class="metric-value">¥{safe_int(cf_financing):,}</span>
                        <span class="metric-change metric-down">▼ 借入返済</span>
                    </div>
                </div>
                <div class="metric-row">
                    <div class="metric-name" style="font-weight: 600;">現金増減</div>
                    <div>
                        <span class="metric-value" style="color: {'#10B981' if cf_net_change >= 0 else '#EF4444'}; font-weight: 700;">¥{safe_int(cf_net_change):,}</span>
                    </div>
                </div>
                """, unsafe_allow_html=True)
                
                st.markdown('</div>', unsafe_allow_html=True)
            
            with col2:
                st.markdown('<div class="section-card fade-in">', unsafe_allow_html=True)
                st.markdown('<h3 class="section-title">💡 推奨アクション</h3>', unsafe_allow_html=True)
                
                st.markdown("""
                <div style="padding: 12px 0;">
                    <div style="margin-bottom: 16px;">
                        <div style="font-weight: 600; color: #1F2937; margin-bottom: 8px;">
                            ✅ 優先度: 高
                        </div>
                        <div style="padding-left: 16px; color: #6B7280; font-size: 0.875rem;">
                            □ 売上債権の早期回収<br>
                            &nbsp;&nbsp;&nbsp;→ 請求サイトの短縮交渉
                        </div>
                    </div>
                    <div style="margin-bottom: 16px;">
                        <div style="font-weight: 600; color: #1F2937; margin-bottom: 8px;">
                            ⚠️ 優先度: 中
                        </div>
                        <div style="padding-left: 16px; color: #6B7280; font-size: 0.875rem;">
                            □ 在庫の適正化<br>
                            &nbsp;&nbsp;&nbsp;→ 滞留在庫の処分検討
                        </div>
                    </div>
                    <div>
                        <div style="font-weight: 600; color: #1F2937; margin-bottom: 8px;">
                            💼 優先度: 中
                        </div>
                        <div style="padding-left: 16px; color: #6B7280; font-size: 0.875rem;">
                            □ 追加融資の検討<br>
                            &nbsp;&nbsp;&nbsp;→ 銀行折衝資料の準備
                        </div>
                    </div>
                </div>
                """, unsafe_allow_html=True)
                
                st.markdown('</div>', unsafe_allow_html=True)
            
            st.markdown("<div style='height: 24px;'></div>", unsafe_allow_html=True)
            
            # データ状態の表示
            if use_real_data:
                st.success("""
                ✅ **実データで動作中！**
                - BS・PLデータから自動でCF計算
                - 実績データに基づく予測
                - リアルタイムアラート表示
                """)
            else:
                st.info("""
                💡 **サンプルデータ表示中**
                
                実データを使用するには：
                1. 「データ取込」メニューを開く
                2. 「BS・CFインポート」タブを選択
                3. BS・PLを含むExcelファイルをアップロード
                
                ファイル要件：
                - シート「貸･事業所(合計)」（BS）
                - シート「損･事業所(合計)」（PL）
                """)

        elif st.session_state.page == "損益計算書 (PL)":
            st.title("損益計算書 (PL)")
            
            st.markdown(f"""
            <div class="info-box">
                <strong>🏢 {st.session_state.selected_comp_name}</strong> | 
                第{st.session_state.selected_period_num}期 | 
                実績締月: {st.session_state.current_month} | 
                シナリオ: <strong>{st.session_state.scenario}</strong>
            </div>
            """, unsafe_allow_html=True)
            
            # PLデータ取得
            if 'pl_df' in st.session_state and st.session_state.pl_df is not None:
                pl_df = st.session_state.pl_df
                
                # 表示モードでフィルタ
                if st.session_state.get('display_mode') == "要約":
                    display_df = pl_df[pl_df['タイプ'] == '要約'].copy()
                else:
                    display_df = pl_df.copy()
            else:
                display_df = pd.DataFrame()
            
            # フィルタリング
            col1, col2 = st.columns([2, 1])
            with col1:
                search_term = st.text_input("🔍 項目名で検索", "")
            
            if not display_df.empty and search_term:
                display_df = display_df[display_df['項目名'].str.contains(search_term, na=False)]
            
            if not display_df.empty:
                # フォーマット
                formatted_df = display_df.style\
                    .format(lambda x: f"¥{safe_int(x):,}" if isinstance(x, (int, float)) else x)\
                    .apply(lambda row: ['background-color: #f8f9fa; font-weight: bold' if row.get('タイプ') == '要約' else '' for _ in row], axis=1)
                
                st.dataframe(formatted_df, width="stretch", height=700)
                
                # CSVダウンロード
                csv = display_df.to_csv(index=False).encode('utf-8-sig')
                st.download_button(
                    "📥 CSVとしてダウンロード",
                    csv,
                    f"PL_{st.session_state.selected_comp_name}_第{st.session_state.selected_period_num}期.csv",
                    "text/csv",
                    key='download-csv'
                )
                
                # ウォーターフォール
                st.markdown("---")
                st.subheader("📊 ウォーターフォールチャート")
                
                # デバッグ情報
                with st.expander("🔍 ウォーターフォールデバッグ", expanded=False):
                    st.write("display_df shape:", display_df.shape)
                    st.write("Columns:", display_df.columns.tolist())
                    if '項目名' in display_df.columns:
                        st.write("項目名:", display_df['項目名'].tolist()[:10])
                    if '金額' in display_df.columns:
                        st.write("金額サンプル:", display_df['金額'].head().tolist())
                    st.dataframe(display_df.head(10))
                
                fig = create_pl_waterfall(display_df)
                if fig:
                    st.plotly_chart(fig, width="stretch")
                else:
                    st.info("ウォーターフォールを表示するにはデータが必要です")
            else:
                st.info("PLデータがありません。実績データまたは予測データを入力してください。")


        # 貸借対照表 (BS) ページ
        elif st.session_state.page == "貸借対照表 (BS)":
            st.title("貸借対照表 (BS)")
            
            st.markdown(f"""
            <div class="info-box">
                <strong>🏢 {st.session_state.selected_comp_name}</strong> | 
                第{st.session_state.selected_period_num}期 | 
                実績締月: {st.session_state.current_month} | 
                シナリオ: <strong>{st.session_state.scenario}</strong>
            </div>
            """, unsafe_allow_html=True)
            
            # デバッグ情報
            with st.expander("🔍 BSデバッグ情報", expanded=False):
                st.write("**データ取得確認:**")
                st.write("selected_period_id:", selected_period_id)
            
            # BSデータ取得 - actual_dataから直接全データを取得
            try:
                # データベースから直接全項目を取得（all_itemsフィルタなし）
                query = """
                SELECT DISTINCT item_name as 項目名, month, amount 
                FROM actual_data 
                WHERE fiscal_period_id = ?
                ORDER BY item_name, month
                """
                
                all_data_df = processor._read_sql_query(query, params=(selected_period_id,))
                
                st.write("**データベース直接取得:**")
                st.write("all_data_df loaded:", all_data_df is not None)
                
                if all_data_df is not None and not all_data_df.empty:
                    st.write("Shape:", all_data_df.shape)
                    st.write("Unique items:", all_data_df['項目名'].nunique())
                    
                    # 横持ち変換
                    pivot_df = all_data_df.pivot(index='項目名', columns='month', values='amount').reset_index()
                    st.write("Pivot shape:", pivot_df.shape)
                    
                    # 縦持ちに変換
                    actuals_df = convert_wide_to_long(pivot_df)
                    st.write("After long conversion:", actuals_df.shape)
                    
                    # カラム名正規化
                    actuals_df.columns = actuals_df.columns.str.lower()
                    
                    # 現在月フィルタ
                    current_month_str = str(st.session_state.current_month)
                    if '-' in current_month_str:
                        current_month_num = int(current_month_str.split('-')[-1])
                    else:
                        current_month_num = int(current_month_str)
                    
                    month_col = None
                    for col in ['fiscal_month', 'month', '月']:
                        if col in actuals_df.columns:
                            month_col = col
                            break
                    
                    if month_col:
                        actuals_df = actuals_df[actuals_df[month_col] == current_month_num]
                        st.write(f"Current month {current_month_num} data:", actuals_df.shape)
                    
                    # カラム検出
                    account_col = None
                    for col in ['account_name', '科目', 'item_name', '項目名']:
                        if col in actuals_df.columns:
                            account_col = col
                            break
                    
                    amount_col = None
                    for col in ['amount', '金額', 'value']:
                        if col in actuals_df.columns:
                            amount_col = col
                            break
                    
                    st.write(f"Detected: account={account_col}, amount={amount_col}")
                    
                    if account_col and amount_col:
                        # 全科目表示
                        current_accounts = actuals_df[account_col].unique().tolist()
                        st.write(f"**当月の科目 ({len(current_accounts)}):**")
                        st.write(current_accounts)
                        
                        # BS科目キーワード
                        bs_keywords = [
                            '資産', '負債', '純資産', '資本',
                            '現金', '預金', '売掛', '買掛', '借入',
                            '棚卸', '在庫', '固定', '投資', '引当'
                        ]
                        
                        # BS科目抽出
                        bs_df = actuals_df[
                            actuals_df[account_col].astype(str).str.contains(
                                '|'.join(bs_keywords),
                                na=False,
                                case=False
                            )
                        ].copy()
                        
                        st.write(f"**BS items found: {len(bs_df)}**")
                        
                        if not bs_df.empty:
                            bs_df = bs_df.rename(columns={account_col: '項目名', amount_col: '金額'})
                            
                            st.success(f"✅ {len(bs_df)}件のBS科目が見つかりました！")
                            
                            # データ表示
                            st.dataframe(
                                bs_df.style.format({'金額': lambda x: f"¥{safe_int(x):,}"}),
                                width="stretch",
                                height=400
                            )
                            
                            # Sankey図
                            st.markdown("---")
                            st.subheader("📊 資金の流れ (Sankey)")
                            
                            fig = create_bs_sankey(bs_df)
                            if fig:
                                st.plotly_chart(fig, width="stretch")
                            else:
                                st.info("Sankey図を表示するにはデータが必要です")
                        else:
                            st.warning("⚠️ BS科目が見つかりませんでした")
                            st.info("BSデータを表示するにはBS科目を含むデータをインポートしてください。")
                    else:
                        st.error("必要なカラムが見つかりません")
                else:
                    st.info("実績データを入力してください")
            except Exception as e:
                st.error(f"エラー: {e}")
                import traceback
                st.code(traceback.format_exc())


        elif st.session_state.page == "予測データ入力":
            st.title("月次計画（予測入力）")
            
            # ヘッダー情報
            col1, col2, col3 = st.columns([2, 2, 2])
            with col1:
                st.markdown(f"**シナリオ:** {st.session_state.scenario}")
            with col2:
                st.markdown(f"**実績締月:** {st.session_state.current_month}")
            with col3:
                st.markdown(f"**期間:** {st.session_state.selected_period_num}期")
            
            st.markdown("---")
            
            # 一括入力機能（Manageboard風）
            with st.expander("🔧 入力アシスト機能", expanded=False):
                st.markdown("#### 一括入力・コピー機能")
            
            st.markdown("---")
            
            # シナリオ自動生成
            st.subheader("🎯 シナリオ自動生成")
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.info("**標準シナリオ**\n現在の入力値をそのまま使用")
            
            with col2:
                st.success("**楽観シナリオ**\n- 売上高: +10%\n- 売上原価: -10%\n- 販管費: -5%")
                if st.button("楽観シナリオを生成", width="stretch"):
                    st.info("楽観シナリオ生成機能は次のバージョンで実装予定です")
            
            with col3:
                st.error("**悲観シナリオ**\n- 売上高: -10%\n- 売上原価: +10%\n- 販管費: +10%")
                if st.button("悲観シナリオを生成", width="stretch"):
                    st.info("悲観シナリオ生成機能は次のバージョンで実装予定です")
            
            st.markdown("---")

            
            st.markdown("---")
            
            # 自動予測機能
            st.subheader("🤖 AI自動予測からインポート")
            
            col1, col2 = st.columns([2, 1])
            with col1:
                st.info("💡 過去データから自動予測した値を一括インポートできます")
            with col2:
                if st.button("🔮 AI予測を実行", type="primary", width="stretch"):
                    st.session_state.page = "AI自動予測"
                    st.rerun()

                
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown("**1. 前期実績をコピー**")
                    if st.button("📋 前期実績をコピー", width="stretch"):
                        # 前期のデータを現在のシナリオにコピー
                        st.info("前期実績コピー機能は今後実装予定です")
                
                with col2:
                    st.markdown("**2. 一括入力（毎月同額）**")
                    
                    # 項目選択
                    editable_items_list = [item for item in processor.all_items if item not in processor.calculated_items]
                    selected_item = st.selectbox(
                        "項目を選択",
                        editable_items_list,
                        key="bulk_input_item"
                    )
                    
                    # 金額入力
                    bulk_amount = st.number_input(
                        "毎月の金額",
                        value=0,
                        step=1000,
                        key="bulk_amount"
                    )
                    
                    if st.button("✏️ 全月に適用", width="stretch", key="apply_bulk"):
                        if bulk_amount != 0:
                            # 全月に同じ金額を設定
                            values = {month: bulk_amount for month in months}
                            success, msg = processor.save_forecast_item(
                                st.session_state.selected_period_id,
                                st.session_state.scenario,
                                selected_item,
                                values
                            )
                            if success:
                                st.success(f"✅ {selected_item}に全月{bulk_amount:,}千円を設定しました")
                                # キャッシュを完全にクリア
                                st.cache_data.clear()
                                for key in ['forecasts_df', 'forecast_data_cache', 'pl_df', 'sub_accounts_df', 
                                           'actuals_df', 'sub_account_aggregation_cache', 'forecast_input_cache_key',
                                           'forecast_input_data']:
                                    if key in st.session_state:
                                        del st.session_state[key]
                                st.rerun()
                            else:
                                st.error(f"❌ {msg}")
                        else:
                            st.warning("⚠️ 金額を入力してください")
                
                st.markdown("---")
                
                # 前年比率での計算
                col1, col2 = st.columns(2)
                with col1:
                    st.markdown("**3. 前年×係数で計算**")
                    ratio = st.number_input(
                        "係数（例: 1.1 = 前年の110%）",
                        value=1.0,
                        step=0.1,
                        min_value=0.1,
                        max_value=10.0,
                        key="ratio_input"
                    )
                    
                    if st.button("🔢 前年×係数で計算", width="stretch"):
                        st.info("前年比率計算機能は今後実装予定です")
            
            st.markdown("---")
            
            # 予測データと補助科目データを取得（現在のシナリオ）
            # キャッシュをチェックして、なければ再読み込み
            cache_key = (st.session_state.selected_period_id, st.session_state.scenario)
            if 'forecast_input_cache_key' not in st.session_state or st.session_state.forecast_input_cache_key != cache_key:
                # 最新データを読み込み
                forecast_data = load_forecast_data_cached(
                    st.session_state.selected_period_id,
                    st.session_state.scenario,
                    processor
                )
                st.session_state.forecast_input_cache_key = cache_key
                st.session_state.forecast_input_data = forecast_data
            else:
                forecast_data = st.session_state.forecast_input_data.copy()
            
            sub_accounts_data = load_sub_accounts_cached(
                st.session_state.selected_period_id,
                st.session_state.scenario,
                processor
            )
            
            # 編集可能な全項目のリストを作成
            editable_items = [item for item in processor.all_items if item not in processor.calculated_items]
            
            # テーブルデータを構築
            table_rows = []
            
            for item in editable_items:
                # 補助科目がある場合、親項目の値を補助科目の合計に置き換える
                if item in processor.parent_items_with_sub_accounts:
                    item_subs = sub_accounts_data[sub_accounts_data['parent_item'] == item]
                    
                    if not item_subs.empty:
                        # 親項目の行（補助科目の合計を表示）
                        item_row = {"項目名": item, "タイプ": "基本", "親項目": item}
                        
                        for month in months:
                            # 補助科目の合計を計算
                            month_subs = item_subs[item_subs['month'] == month]
                            if not month_subs.empty:
                                total = month_subs['amount'].sum()
                                item_row[month] = float(total)
                            else:
                                item_row[month] = 0.0
                        
                        table_rows.append(item_row)
                        
                        # 補助科目の行
                        for sub_name in item_subs['sub_account_name'].unique():
                            sub_row = {"項目名": f"  └ {sub_name}", "タイプ": "補助", "親項目": item}
                            sub_data = item_subs[item_subs['sub_account_name'] == sub_name]
                            
                            for month in months:
                                month_data = sub_data[sub_data['month'] == month]
                                if not month_data.empty:
                                    val = month_data['amount'].iloc[0]
                                    sub_row[month] = float(val) if pd.notna(val) else 0.0
                                else:
                                    sub_row[month] = 0.0
                            
                            table_rows.append(sub_row)
                    else:
                        # 補助科目がない場合は通常通り
                        item_row = {"項目名": item, "タイプ": "基本", "親項目": item}
                        item_data = forecast_data[forecast_data['項目名'] == item]
                        
                        for month in months:
                            if not item_data.empty and month in item_data.columns:
                                val = item_data[month].iloc[0]
                                item_row[month] = float(val) if pd.notna(val) else 0.0
                            else:
                                item_row[month] = 0.0
                        
                        table_rows.append(item_row)
                else:
                    # 補助科目を持たない項目は通常通り
                    item_row = {"項目名": item, "タイプ": "基本", "親項目": item}
                    item_data = forecast_data[forecast_data['項目名'] == item]
                    
                    for month in months:
                        if not item_data.empty and month in item_data.columns:
                            val = item_data[month].iloc[0]
                            item_row[month] = float(val) if pd.notna(val) else 0.0
                        else:
                            item_row[month] = 0.0
                    
                    table_rows.append(item_row)
            
            # DataFrameに変換
            edit_df = pd.DataFrame(table_rows)
            
            # 合計列を追加
            month_cols = [m for m in months if m in edit_df.columns]
            edit_df['合計'] = edit_df[month_cols].sum(axis=1)
            
            # カラム設定（カンマ区切り）
            column_config = {
                "項目名": st.column_config.TextColumn("項目名", width="large", disabled=True),
                "タイプ": st.column_config.TextColumn("タイプ", width="small", disabled=True),
                "親項目": None,  # 非表示
                "合計": st.column_config.NumberColumn("合計", format="%.0f", disabled=True, width="medium")
            }
            
            for month in month_cols:
                column_config[month] = st.column_config.NumberColumn(
                    month,
                    format="%.0f",  # カンマ区切り（通貨記号なし）
                    width="small",
                    help=f"{month}の予測値（千円）"
                )
            
            # データエディタで全体を表示・編集
            st.markdown("### 予測損益計算書（スプレッドシート）")
            
            # 補助科目がある親項目のリストを作成
            items_with_subs = []
            for item in editable_items:
                if item in processor.parent_items_with_sub_accounts:
                    item_subs = sub_accounts_data[sub_accounts_data['parent_item'] == item]
                    if not item_subs.empty:
                        items_with_subs.append(item)
            
            # デバッグ情報（開発時のみ）
            if st.checkbox("🔧 デバッグ情報を表示", value=False, key="debug_forecast_input"):
                st.write("データ件数:", len(forecast_data))
                st.write("補助科目データ件数:", len(sub_accounts_data))
                st.write("表示行数:", len(edit_df))
                st.write("編集可能項目数:", len(editable_items))
                st.write("補助科目がある項目:", items_with_subs)
                st.write("月列:", month_cols)
            
            st.markdown("💡 数値をクリックして直接編集できます。補助科目がある項目は自動計算されます。")
            
            # 項目名列と合計列、タイプ列は常に編集不可
            disabled_columns = ["項目名", "タイプ", "合計"]
            
            edited_df = st.data_editor(
                edit_df,
                column_config=column_config,
                width="stretch",
                height=600,
                key="forecast_pl_editor",
                hide_index=True,
                disabled=disabled_columns,  # 項目名、タイプ、合計列のみ編集不可
                num_rows="fixed"  # 行の追加・削除を禁止
            )
            
            # 保存ボタン
            col1, col2, col3 = st.columns([2, 2, 1])
            
            with col1:
                if st.button("💾 すべての変更を保存", type="primary", key="save_all_forecast"):
                    with st.spinner("保存中..."):
                        success_count = 0
                        error_count = 0
                        
                        # 警告メッセージ
                        if items_with_subs:
                            st.info(f"ℹ️ 補助科目がある項目（{', '.join(items_with_subs)}）は自動計算されるため、保存時にスキップされます。")
                        
                        # 基本項目を保存（補助科目がない項目のみ）
                        for _, row in edited_df[edited_df['タイプ'] == '基本'].iterrows():
                            item_name = row['項目名']
                            
                            # 補助科目がある項目はスキップ（自動計算されるため）
                            if item_name in items_with_subs:
                                continue
                            
                            values = {month: row[month] for month in month_cols}
                            
                            success, msg = processor.save_forecast_item(
                                st.session_state.selected_period_id,
                                st.session_state.scenario,
                                item_name,
                                values
                            )
                            
                            if success:
                                success_count += 1
                            else:
                                error_count += 1
                                st.error(f"❌ {item_name}: {msg}")
                        
                        # 補助科目を保存
                        for _, row in edited_df[edited_df['タイプ'] == '補助'].iterrows():
                            full_name = row['項目名']
                            sub_name = full_name.replace('  └ ', '')
                            parent_item = row['親項目']  # 親項目情報を直接取得
                            
                            values = {month: row[month] for month in month_cols}
                            
                            success, msg = processor.save_sub_account(
                                st.session_state.selected_period_id,
                                st.session_state.scenario,
                                parent_item,
                                sub_name,
                                values
                            )
                            
                            if success:
                                success_count += 1
                            else:
                                error_count += 1
                                st.error(f"❌ {sub_name}: {msg}")
                        
                        if error_count == 0:
                            st.success(f"✅ {success_count}件のデータを保存しました")
                            # キャッシュクリア
                            st.cache_data.clear()
                            for key in ['forecasts_df', 'sub_accounts_df', 'pl_df', 'forecast_input_cache_key', 'forecast_input_data']:
                                if key in st.session_state:
                                    del st.session_state[key]
                            st.rerun()
                        else:
                            st.warning(f"⚠️ {success_count}件成功、{error_count}件失敗")
            
            with col2:
                # 補助科目の追加機能
                with st.expander("➕ 補助科目を追加"):
                    parent_item = st.selectbox(
                        "親項目を選択",
                        processor.parent_items_with_sub_accounts,
                        key="add_sub_parent"
                    )
                    
                    new_sub_name = st.text_input(
                        "補助科目名",
                        key="add_sub_name",
                        placeholder="例: 国内売上"
                    )
                    
                    if new_sub_name and st.button("追加", key="add_sub_confirm"):
                        # 初期値はすべて0
                        values = {month: 0.0 for month in months}
                        
                        success, msg = processor.save_sub_account(
                            st.session_state.selected_period_id,
                            st.session_state.scenario,
                            parent_item,
                            new_sub_name,
                            values
                        )
                        
                        if success:
                            st.success(f"✅ {new_sub_name}を追加しました")
                            st.cache_data.clear()
                            st.rerun()
                        else:
                            st.error(f"❌ {msg}")
            
            with col3:
                if st.button("🔄 リセット", key="reset_forecast"):
                    st.cache_data.clear()
                    if 'forecasts_df' in st.session_state:
                        del st.session_state.forecasts_df
                    if 'sub_accounts_df' in st.session_state:
                        del st.session_state.sub_accounts_df
                    st.rerun()
            
        
        
        elif st.session_state.page == "キャッシュフロー計算書 (CF)":
            st.title("キャッシュフロー計算書")
            
            st.markdown("""
            <div class="info-box">
                <strong>💡 概要:</strong> 資金の流れを「営業活動」「投資活動」「財務活動」に分けて把握します。
            </div>
            """, unsafe_allow_html=True)
            
            # CFデータを計算
            cf_data = processor.calculate_cash_flow(st.session_state.selected_period_id)
            
            if cf_data:
                # 各カテゴリのサマリーカード
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    operating_cf_total = sum([v for v in cf_data.get("営業活動によるキャッシュフロー", {}).values() if pd.notna(v)])
                    st.markdown(f"""
                    <div class="summary-card-blue">
                        <div class="card-title">営業活動CF</div>
                        <div class="card-value">¥{safe_int(operating_cf_total):,}</div>
                        <div class="card-subtitle">本業で稼いだ現金</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col2:
                    investing_cf_total = sum([v for v in cf_data.get("投資活動によるキャッシュフロー", {}).values() if pd.notna(v)])
                    st.markdown(f"""
                    <div class="summary-card-orange">
                        <div class="card-title">投資活動CF</div>
                        <div class="card-value">¥{safe_int(investing_cf_total):,}</div>
                        <div class="card-subtitle">設備投資などの支出</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col3:
                    financing_cf_total = sum([v for v in cf_data.get("財務活動によるキャッシュフロー", {}).values() if pd.notna(v)])
                    st.markdown(f"""
                    <div class="summary-card-purple">
                        <div class="card-title">財務活動CF</div>
                        <div class="card-value">¥{safe_int(financing_cf_total):,}</div>
                        <div class="card-subtitle">借入・返済の収支</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                # 詳細テーブル
                st.markdown("### 📊 月次キャッシュフロー推移")
                
                cf_rows = []
                for category, month_data in cf_data.items():
                    row = {"項目": category}
                    row.update(month_data)
                    cf_rows.append(row)
                
                if cf_rows:
                    cf_df = pd.DataFrame(cf_rows)
                    st.dataframe(cf_df, width="stretch", height=300)
                    
                    # グラフ
                    st.markdown("### 📈 キャッシュフロー推移グラフ")
                    fig = go.Figure()
                    
                    for category in cf_data.keys():
                        months_list = list(cf_data[category].keys())
                        values_list = [cf_data[category][m] for m in months_list]
                        
                        fig.add_trace(go.Scatter(
                            x=months_list,
                            y=values_list,
                            mode='lines+markers',
                            name=category,
                            line=dict(width=3)
                        ))
                    
                    fig.update_layout(
                        xaxis_title="月",
                        yaxis_title="金額 (円)",
                        hovermode='x unified',
                        height=500,
                        template="plotly_white"
                    )
                    
                    st.plotly_chart(fig, width="stretch")
            else:
                st.warning("キャッシュフローデータがありません。")
        
        elif st.session_state.page == "CF詳細分析":
            st.title("💰 キャッシュフロー詳細分析")
            
            # データチェック
            if 'cf_data' not in st.session_state or not st.session_state.cf_data:
                st.warning("⚠️ CFデータが読み込まれていません")
                st.info("「データ取込」→「BS・CFインポート」からデータをアップロードしてください")
            else:
                cf_data = st.session_state.cf_data
                
                st.success(f"✅ {len(cf_data)}ヶ月分のCFデータを読み込み済み")
                
                # 月次CF計算書詳細
                st.markdown("### 📊 月次キャッシュフロー計算書")
                
                # 月選択
                available_months = list(cf_data.keys())
                selected_month = st.selectbox(
                    "月を選択",
                    available_months,
                    index=len(available_months)-1 if available_months else 0
                )
                
                if selected_month:
                    cf_month = cf_data[selected_month]
                    
                    # 3カラムで表示
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.markdown('<div class="section-card">', unsafe_allow_html=True)
                        st.markdown("#### 🔄 営業活動によるCF")
                        
                        operating_cf = cf_month.get('営業CF', {})
                        st.markdown(f"""
                        | 項目 | 金額 |
                        |------|------|
                        | 税引前利益 | ¥{safe_int(operating_cf.get('税引前利益', 0)):,} |
                        | + 減価償却費 | ¥{safe_int(operating_cf.get('減価償却費', 0)):,} |
                        | - 売上債権増加 | ¥{safe_int(operating_cf.get('売上債権の増減', 0)):,} |
                        | - 棚卸資産増加 | ¥{safe_int(operating_cf.get('棚卸資産の増減', 0)):,} |
                        | + 買入債務増加 | ¥{safe_int(operating_cf.get('買入債務の増減', 0)):,} |
                        | - 法人税支払 | ¥{safe_int(operating_cf.get('法人税の支払', 0)):,} |
                        | **= 営業CF** | **¥{safe_int(operating_cf.get('合計', 0)):,}** |
                        """)
                        st.markdown('</div>', unsafe_allow_html=True)
                    
                    with col2:
                        st.markdown('<div class="section-card">', unsafe_allow_html=True)
                        st.markdown("#### 📈 投資活動によるCF")
                        
                        investing_cf = cf_month.get('投資CF', {})
                        st.markdown(f"""
                        | 項目 | 金額 |
                        |------|------|
                        | - 固定資産取得 | ¥{safe_int(investing_cf.get('固定資産の取得', 0)):,} |
                        | **= 投資CF** | **¥{safe_int(investing_cf.get('合計', 0)):,}** |
                        """)
                        st.markdown('</div>', unsafe_allow_html=True)
                    
                    with col3:
                        st.markdown('<div class="section-card">', unsafe_allow_html=True)
                        st.markdown("#### 💼 財務活動によるCF")
                        
                        financing_cf = cf_month.get('財務CF', {})
                        st.markdown(f"""
                        | 項目 | 金額 |
                        |------|------|
                        | - 借入金返済 | ¥{safe_int(financing_cf.get('借入金の返済', 0)):,} |
                        | - 配当金支払 | ¥{safe_int(financing_cf.get('配当金の支払', 0)):,} |
                        | **= 財務CF** | **¥{safe_int(financing_cf.get('合計', 0)):,}** |
                        """)
                        st.markdown('</div>', unsafe_allow_html=True)
                    
                    # 現金増減サマリー
                    st.markdown("---")
                    st.markdown("### 💵 現金増減サマリー")
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        st.metric(
                            "営業CF + 投資CF + 財務CF",
                            f"¥{safe_int(cf_month.get('現金増減', 0)):,}",
                            delta=None
                        )
                    
                    with col2:
                        st.metric(
                            "期末現金残高",
                            f"¥{safe_int(cf_month.get('期末現金', 0)):,}",
                            delta=f"¥{safe_int(cf_month.get('現金増減', 0)):,}"
                        )
                    
                    # 累積CF推移グラフ
                    st.markdown("---")
                    st.markdown("### 📊 累積キャッシュフロー推移")
                    
                    # データ準備
                    months_list = []
                    cumulative_operating = []
                    cumulative_investing = []
                    cumulative_financing = []
                    cumulative_total = []
                    
                    running_operating = 0
                    running_investing = 0
                    running_financing = 0
                    running_total = 0
                    
                    for month_key in cf_data.keys():
                        cf_m = cf_data[month_key]
                        months_list.append(month_key.replace('月度', '月'))
                        
                        running_operating += cf_m.get('営業CF', {}).get('合計', 0)
                        running_investing += cf_m.get('投資CF', {}).get('合計', 0)
                        running_financing += cf_m.get('財務CF', {}).get('合計', 0)
                        running_total += cf_m.get('現金増減', 0)
                        
                        cumulative_operating.append(running_operating)
                        cumulative_investing.append(running_investing)
                        cumulative_financing.append(running_financing)
                        cumulative_total.append(running_total)
                    
                    fig = go.Figure()
                    
                    fig.add_trace(go.Scatter(
                        x=months_list, y=cumulative_operating,
                        name='営業CF（累積）',
                        line=dict(color='#10B981', width=3)
                    ))
                    
                    fig.add_trace(go.Scatter(
                        x=months_list, y=cumulative_investing,
                        name='投資CF（累積）',
                        line=dict(color='#EF4444', width=3)
                    ))
                    
                    fig.add_trace(go.Scatter(
                        x=months_list, y=cumulative_financing,
                        name='財務CF（累積）',
                        line=dict(color='#F59E0B', width=3)
                    ))
                    
                    fig.add_trace(go.Scatter(
                        x=months_list, y=cumulative_total,
                        name='現金増減（累積）',
                        line=dict(color='#3B82F6', width=4)
                    ))
                    
                    fig.update_layout(
                        template='plotly_white',
                        hovermode='x unified',
                        height=400,
                        yaxis=dict(tickformat=',.0f')
                    )
                    
                    st.plotly_chart(fig, width="stretch")
        
        elif st.session_state.page == "運転資本分析":
            st.title("🔄 運転資本分析")
            
            # データチェック
            if 'bs_data' not in st.session_state or st.session_state.bs_data.empty:
                st.warning("⚠️ BSデータが読み込まれていません")
                st.info("「データ取込」→「BS・CFインポート」からデータをアップロードしてください")
            else:
                bs_data = st.session_state.bs_data
                
                st.success(f"✅ BSデータを読み込み済み")
                
                # 運転資本指標を計算
                with st.spinner("💰 運転資本指標を計算しています..."):
                    wc_metrics = cf_analyzer.calculate_working_capital_metrics(
                        bs_data,
                        None  # PLデータ
                    )
                
                if wc_metrics:
                    # 最新月の指標を表示
                    latest_month = list(wc_metrics.keys())[-1]
                    latest_metrics = wc_metrics[latest_month]
                    
                    st.markdown("### 📊 運転資本KPI（最新月）")
                    
                    col1, col2, col3, col4 = st.columns(4)
                    
                    with col1:
                        st.metric(
                            "運転資本",
                            f"¥{safe_int(latest_metrics['運転資本']):,}",
                            help="(売上債権 + 棚卸資産) - 買入債務"
                        )
                    
                    with col2:
                        st.metric(
                            "売上債権回転日数",
                            f"{latest_metrics['売上債権回転日数']:.1f}日",
                            help="売上債権 ÷ 月間売上 × 30"
                        )
                    
                    with col3:
                        st.metric(
                            "在庫回転日数",
                            f"{latest_metrics['棚卸資産回転日数']:.1f}日",
                            help="棚卸資産 ÷ 月間売上原価 × 30"
                        )
                    
                    with col4:
                        st.metric(
                            "CCC",
                            f"{latest_metrics['CCC']:.1f}日",
                            help="キャッシュコンバージョンサイクル"
                        )
                    
                    st.markdown("---")
                    
                    # 運転資本の内訳
                    st.markdown("### 🔍 運転資本の内訳")
                    
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.markdown('<div class="section-card">', unsafe_allow_html=True)
                        st.markdown("#### 売上債権")
                        st.metric("", f"¥{safe_int(latest_metrics['売上債権']):,}")
                        st.markdown('</div>', unsafe_allow_html=True)
                    
                    with col2:
                        st.markdown('<div class="section-card">', unsafe_allow_html=True)
                        st.markdown("#### 棚卸資産")
                        st.metric("", f"¥{safe_int(latest_metrics['棚卸資産']):,}")
                        st.markdown('</div>', unsafe_allow_html=True)
                    
                    with col3:
                        st.markdown('<div class="section-card">', unsafe_allow_html=True)
                        st.markdown("#### 買入債務")
                        st.metric("", f"¥{safe_int(latest_metrics['買入債務']):,}")
                        st.markdown('</div>', unsafe_allow_html=True)
                    
                    st.markdown("---")
                    
                    # 推移グラフ
                    st.markdown("### 📈 運転資本の推移")
                    
                    months_list = []
                    wc_values = []
                    ar_values = []
                    inv_values = []
                    ap_values = []
                    
                    for month, metrics in wc_metrics.items():
                        months_list.append(month)
                        wc_values.append(metrics['運転資本'])
                        ar_values.append(metrics['売上債権'])
                        inv_values.append(metrics['棚卸資産'])
                        ap_values.append(metrics['買入債務'])
                    
                    fig = go.Figure()
                    
                    fig.add_trace(go.Scatter(
                        x=months_list, y=wc_values,
                        name='運転資本',
                        line=dict(color='#3B82F6', width=4),
                        fill='tozeroy'
                    ))
                    
                    fig.update_layout(
                        template='plotly_white',
                        hovermode='x',
                        height=400,
                        yaxis=dict(tickformat=',.0f', title="金額（円）")
                    )
                    
                    st.plotly_chart(fig, width="stretch")
                    
                    # CCC推移
                    st.markdown("### ⏱️ CCC（キャッシュコンバージョンサイクル）推移")
                    
                    ccc_values = [m['CCC'] for m in wc_metrics.values()]
                    
                    fig = go.Figure()
                    
                    fig.add_trace(go.Scatter(
                        x=months_list, y=ccc_values,
                        name='CCC',
                        line=dict(color='#F59E0B', width=3),
                        mode='lines+markers'
                    ))
                    
                    fig.update_layout(
                        template='plotly_white',
                        hovermode='x',
                        height=350,
                        yaxis=dict(title="日数")
                    )
                    
                    st.plotly_chart(fig, width="stretch")
                    
                    # 分析コメント
                    st.markdown("---")
                    st.markdown("### 💡 分析結果")
                    
                    ccc = latest_metrics['CCC']
                    
                    if ccc < 30:
                        st.success(f"✅ CCCは{ccc:.1f}日と短く、効率的な資金運用ができています")
                    elif ccc < 60:
                        st.info(f"💡 CCCは{ccc:.1f}日です。改善の余地があります")
                    else:
                        st.warning(f"⚠️ CCCが{ccc:.1f}日と長期化しています。運転資本の改善が必要です")
                    
                    st.markdown("""
                    **改善施策:**
                    - 売上債権回転日数の短縮 → 請求サイトの見直し
                    - 在庫回転日数の短縮 → 在庫管理の最適化
                    - 買入債務回転日数の延長 → 支払条件の交渉
                    """)
                else:
                    st.error("運転資本指標の計算に失敗しました")
        
        elif st.session_state.page == "経営指標ダッシュボード":
            st.title("経営指標")
            
            st.markdown("""
            <div class="info-box">
                <strong>💡 概要:</strong> 企業の収益性、効率性、安全性を数値で評価します。
            </div>
            """, unsafe_allow_html=True)
            
            # 経営指標を計算
            indicators = processor.calculate_financial_indicators(st.session_state.selected_period_id)
            
            if indicators:
                # 最新月の指標を取得
                latest_month = list(indicators.keys())[-1] if indicators else None
                
                if latest_month:
                    latest_indicators = indicators[latest_month]
                    
                    # KPIカード
                    st.markdown("### 📈 主要経営指標（最新月）")
                    
                    col1, col2, col3, col4 = st.columns(4)
                    
                    with col1:
                        st.markdown(f"""
                        <div class="kpi-card">
                            <div class="kpi-label">粗利率</div>
                            <div class="kpi-value">{latest_indicators['粗利率']:.1f}%</div>
                            <div class="kpi-change kpi-{'positive' if latest_indicators['粗利率'] > 30 else 'negative'}">
                                {'✓ 良好' if latest_indicators['粗利率'] > 30 else '△ 要改善'}
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    with col2:
                        st.markdown(f"""
                        <div class="kpi-card">
                            <div class="kpi-label">営業利益率</div>
                            <div class="kpi-value">{latest_indicators['営業利益率']:.1f}%</div>
                            <div class="kpi-change kpi-{'positive' if latest_indicators['営業利益率'] > 5 else 'negative'}">
                                {'✓ 良好' if latest_indicators['営業利益率'] > 5 else '△ 要改善'}
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    with col3:
                        st.markdown(f"""
                        <div class="kpi-card">
                            <div class="kpi-label">経常利益率</div>
                            <div class="kpi-value">{latest_indicators['経常利益率']:.1f}%</div>
                            <div class="kpi-change kpi-{'positive' if latest_indicators['経常利益率'] > 3 else 'negative'}">
                                {'✓ 良好' if latest_indicators['経常利益率'] > 3 else '△ 要改善'}
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    with col4:
                        st.markdown(f"""
                        <div class="kpi-card">
                            <div class="kpi-label">当期純利益率</div>
                            <div class="kpi-value">{latest_indicators['当期純利益率']:.1f}%</div>
                            <div class="kpi-change kpi-{'positive' if latest_indicators['当期純利益率'] > 2 else 'negative'}">
                                {'✓ 良好' if latest_indicators['当期純利益率'] > 2 else '△ 要改善'}
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    # 月次推移グラフ
                    st.markdown("### 📈 収益性指標の推移")
                    
                    months_list = list(indicators.keys())
                    
                    fig = go.Figure()
                    
                    fig.add_trace(go.Scatter(
                        x=months_list,
                        y=[indicators[m]['粗利率'] for m in months_list],
                        mode='lines+markers',
                        name='粗利率',
                        line=dict(color='#2e7d32', width=3)
                    ))
                    
                    fig.add_trace(go.Scatter(
                        x=months_list,
                        y=[indicators[m]['営業利益率'] for m in months_list],
                        mode='lines+markers',
                        name='営業利益率',
                        line=dict(color='#1976d2', width=3)
                    ))
                    
                    fig.add_trace(go.Scatter(
                        x=months_list,
                        y=[indicators[m]['経常利益率'] for m in months_list],
                        mode='lines+markers',
                        name='経常利益率',
                        line=dict(color='#f57c00', width=3)
                    ))
                    
                    fig.add_trace(go.Scatter(
                        x=months_list,
                        y=[indicators[m]['当期純利益率'] for m in months_list],
                        mode='lines+markers',
                        name='当期純利益率',
                        line=dict(color='#6a1b9a', width=3)
                    ))
                    
                    fig.update_layout(
                        xaxis_title="月",
                        yaxis_title="利益率 (%)",
                        hovermode='x unified',
                        height=500,
                        template="plotly_white"
                    )
                    
                    st.plotly_chart(fig, width="stretch")
                    
                    # 推奨改善アクション
                    st.markdown("### 💡 推奨改善アクション")
                    
                    if latest_indicators['粗利率'] < 30:
                        st.markdown("""
                        <div class="warning-box">
                            <strong>⚠️ 粗利率が低い</strong><br>
                            • 価格設定の見直し<br>
                            • 原価削減施策の検討<br>
                            • 高付加価値商品へのシフト
                        </div>
                        """, unsafe_allow_html=True)
                    
                    if latest_indicators['営業利益率'] < 5:
                        st.markdown("""
                        <div class="warning-box">
                            <strong>⚠️ 営業利益率が低い</strong><br>
                            • 販管費の見直し<br>
                            • 業務効率化の推進<br>
                            • 固定費の削減検討
                        </div>
                        """, unsafe_allow_html=True)
                    
                    if latest_indicators['経常利益率'] > 3 and latest_indicators['営業利益率'] > 5:
                        st.markdown("""
                        <div class="success-box">
                            <strong>✓ 良好な収益性</strong><br>
                            現在の収益構造を維持しつつ、さらなる成長を目指しましょう。
                        </div>
                        """, unsafe_allow_html=True)
            else:
                st.warning("経営指標データがありません。")
        
        elif st.session_state.page == "損益分岐点分析":
            st.title("損益分岐点分析")
            
            st.markdown("""
            <div class="info-box">
                <strong>💡 概要:</strong> 赤字にならない最低売上高を計算し、経営の安全性を評価します。
            </div>
            """, unsafe_allow_html=True)
            
            # 損益分岐点を計算（予測データを使用）
            forecasts = load_forecast_data_cached(
                st.session_state.selected_period_id,
                st.session_state.scenario,
                processor
            )
            
            if not forecasts.empty:
                months = [col for col in forecasts.columns if col not in ['項目名']]
                
                # 売上高
                sales_row = forecasts[forecasts['項目名'] == '売上高']
                total_sales = 0
                if not sales_row.empty:
                    for month in months:
                        if month in sales_row.columns:
                            val = sales_row[month].iloc[0]
                            if pd.notna(val):
                                total_sales += float(val)
                
                # 変動費（売上原価）
                vc_row = forecasts[forecasts['項目名'] == '売上原価']
                total_vc = 0
                if not vc_row.empty:
                    for month in months:
                        if month in vc_row.columns:
                            val = vc_row[month].iloc[0]
                            if pd.notna(val):
                                total_vc += float(val)
                
                # 固定費（販管費）
                total_fc = 0
                for item in processor.ga_items:
                    item_row = forecasts[forecasts['項目名'] == item]
                    if not item_row.empty:
                        for month in months:
                            if month in item_row.columns:
                                val = item_row[month].iloc[0]
                                if pd.notna(val):
                                    total_fc += float(val)
                
                # 計算
                contribution_margin = total_sales - total_vc
                contribution_margin_ratio = (contribution_margin / total_sales * 100) if total_sales > 0 else 0
                breakeven_sales = (total_fc / (contribution_margin_ratio / 100)) if contribution_margin_ratio > 0 else 0
                safety_margin = total_sales - breakeven_sales
                safety_margin_ratio = (safety_margin / total_sales * 100) if total_sales > 0 else 0
                
                # サマリーカード
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.markdown(f"""
                    <div class="summary-card-blue">
                        <div class="card-title">損益分岐点売上高</div>
                        <div class="card-value">¥{safe_int(breakeven_sales):,}</div>
                        <div class="card-subtitle">この売上で利益ゼロ</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col2:
                    st.markdown(f"""
                    <div class="summary-card-green">
                        <div class="card-title">安全余裕額</div>
                        <div class="card-value">¥{safe_int(safety_margin):,}</div>
                        <div class="card-subtitle">赤字までの余裕</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col3:
                    st.markdown(f"""
                    <div class="summary-card-{'green' if safety_margin_ratio > 20 else 'orange'}">
                        <div class="card-title">安全余裕率</div>
                        <div class="card-value">{safety_margin_ratio:.1f}%</div>
                        <div class="card-subtitle">{'安全' if safety_margin_ratio > 20 else '注意'}</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                # 詳細分析
                st.markdown("### 📊 詳細分析")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown("""
                    <div class="kpi-card">
                        <h4>費用構造</h4>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    analysis_data = {
                        "項目": ["売上高", "変動費", "限界利益", "固定費", "営業利益"],
                        "金額": [total_sales, total_vc, contribution_margin, total_fc, contribution_margin - total_fc],
                        "構成比(%)": [
                            100,
                            (total_vc / total_sales * 100) if total_sales > 0 else 0,
                            (contribution_margin / total_sales * 100) if total_sales > 0 else 0,
                            (total_fc / total_sales * 100) if total_sales > 0 else 0,
                            ((contribution_margin - total_fc) / total_sales * 100) if total_sales > 0 else 0
                        ]
                    }
                    
                    analysis_df = pd.DataFrame(analysis_data)
                    st.dataframe(
                        analysis_df.style.format({"金額": "¥{:,.0f}", "構成比(%)": "{:.1f}%"}),
                        width="stretch"
                    )
                
                with col2:
                    st.markdown("""
                    <div class="kpi-card">
                        <h4>重要指標</h4>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    metrics_data = {
                        "指標": ["限界利益率", "損益分岐点比率", "安全余裕率", "固定費回収率"],
                        "値": [
                            f"{contribution_margin_ratio:.1f}%",
                            f"{(breakeven_sales / total_sales * 100) if total_sales > 0 else 0:.1f}%",
                            f"{safety_margin_ratio:.1f}%",
                            f"{(contribution_margin / total_fc * 100) if total_fc > 0 else 0:.1f}%"
                        ],
                        "評価": [
                            "良好" if contribution_margin_ratio > 40 else "要改善",
                            "良好" if (total_sales > 0 and (breakeven_sales / total_sales * 100) < 80) else ("要改善" if total_sales > 0 else "-"),
                            "良好" if safety_margin_ratio > 20 else "要注意",
                            "良好" if (total_fc > 0 and (contribution_margin / total_fc * 100) > 120) else ("要改善" if total_fc > 0 else "-")
                        ]
                    }
                    
                    metrics_df = pd.DataFrame(metrics_data)
                    st.dataframe(metrics_df, width="stretch")
                
                # グラフ
                st.markdown("### 📈 損益分岐点グラフ")
                
                # X軸（売上高の範囲）
                x_range = np.linspace(0, total_sales * 1.5, 100)
                
                # 総費用線（固定費 + 変動費）
                variable_cost_ratio = total_vc / total_sales if total_sales > 0 else 0
                total_cost_line = total_fc + (x_range * variable_cost_ratio)
                
                # 売上高線
                sales_line = x_range
                
                fig = go.Figure()
                
                # 売上高線
                fig.add_trace(go.Scatter(
                    x=x_range,
                    y=sales_line,
                    mode='lines',
                    name='売上高',
                    line=dict(color='#2e7d32', width=3)
                ))
                
                # 総費用線
                fig.add_trace(go.Scatter(
                    x=x_range,
                    y=total_cost_line,
                    mode='lines',
                    name='総費用',
                    line=dict(color='#c62828', width=3)
                ))
                
                # 損益分岐点
                fig.add_trace(go.Scatter(
                    x=[breakeven_sales],
                    y=[breakeven_sales],
                    mode='markers',
                    name='損益分岐点',
                    marker=dict(size=15, color='#f57c00')
                ))
                
                # 現在の売上
                current_total_cost = total_fc + (total_sales * variable_cost_ratio)
                fig.add_trace(go.Scatter(
                    x=[total_sales],
                    y=[current_total_cost],
                    mode='markers',
                    name='現在位置',
                    marker=dict(size=15, color='#1976d2', symbol='star')
                ))
                
                fig.update_layout(
                    xaxis_title="売上高 (円)",
                    yaxis_title="金額 (円)",
                    hovermode='closest',
                    height=500,
                    template="plotly_white"
                )
                
                st.plotly_chart(fig, width="stretch")
                
                # 改善提案
                st.markdown("### 💡 改善提案")
                
                if safety_margin_ratio < 10:
                    st.markdown("""
                    <div class="warning-box">
                        <strong>⚠️ 危険水準：安全余裕率が10%未満</strong><br>
                        <strong>至急対応が必要：</strong><br>
                        • 固定費の大幅削減を検討<br>
                        • 売上拡大策の即時実行<br>
                        • 変動費率の改善（仕入先交渉など）<br>
                        • 資金繰り計画の見直し
                    </div>
                    """, unsafe_allow_html=True)
                elif safety_margin_ratio < 20:
                    st.markdown("""
                    <div class="warning-box">
                        <strong>⚠️ 注意水準：安全余裕率が20%未満</strong><br>
                        <strong>改善施策：</strong><br>
                        • 固定費の削減余地を調査<br>
                        • 売上増加施策の検討<br>
                        • 利益率の高い商品の強化
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    st.markdown("""
                    <div class="success-box">
                        <strong>✓ 良好な水準：安全余裕率が20%以上</strong><br>
                        現在の経営は安全です。さらなる成長を目指しましょう。
                    </div>
                    """, unsafe_allow_html=True)
            else:
                st.warning("予測データがありません。予測データを入力してください。")
        
        elif st.session_state.page == "予測 VS 実績比較":
            st.title("予実比較")
            
            st.markdown("""
            <div class="info-box">
                <strong>💡 使い方:</strong> 予測値と実績値の差異を分析します。達成率や乖離額を確認できます。
            </div>
            """, unsafe_allow_html=True)
            
            # 実績データと予測データを取得
            actuals = actuals_df.copy()
            forecasts = load_forecast_data_cached(
                st.session_state.selected_period_id,
                st.session_state.scenario,
                processor
            )
            
            # 比較テーブルを作成
            comparison_rows = []
            
            for item in processor.all_items:
                actual_row = actuals[actuals['項目名'] == item]
                forecast_row = forecasts[forecasts['項目名'] == item]
                
                if actual_row.empty or forecast_row.empty:
                    continue
                
                row_data = {"項目名": item}
                
                # 実績合計
                actual_total = 0
                for month in months:
                    if month in actual_row.columns:
                        val = actual_row[month].iloc[0]
                        if pd.notna(val):
                            actual_total += float(val)
                
                # 予測合計
                forecast_total = 0
                for month in months:
                    if month in forecast_row.columns:
                        val = forecast_row[month].iloc[0]
                        if pd.notna(val):
                            forecast_total += float(val)
                
                # 差異計算
                diff = actual_total - forecast_total
                diff_rate = (diff / forecast_total * 100) if forecast_total != 0 else 0
                achievement_rate = (actual_total / forecast_total * 100) if forecast_total != 0 else 0
                
                row_data["実績"] = actual_total
                row_data["予測"] = forecast_total
                row_data["差異"] = diff
                row_data["差異率(%)"] = diff_rate
                row_data["達成率(%)"] = achievement_rate
                
                comparison_rows.append(row_data)
            
            comparison_df = pd.DataFrame(comparison_rows)
            
            # フォーマットして表示
            if not comparison_df.empty:
                formatted_df = comparison_df.style\
                    .format({
                        "実績": "¥{:,.0f}",
                        "予測": "¥{:,.0f}",
                        "差異": "¥{:,.0f}",
                        "差異率(%)": "{:.1f}%",
                        "達成率(%)": "{:.1f}%"
                    })\
                    .applymap(
                        lambda x: 'background-color: #d4edda' if isinstance(x, (int, float)) and x > 0 else 
                                  ('background-color: #f8d7da' if isinstance(x, (int, float)) and x < 0 else ''),
                        subset=['差異', '差異率(%)']
                    )
                
                st.dataframe(formatted_df, width="stretch", height=600)
                
                # CSVダウンロード
                csv = comparison_df.to_csv(index=False).encode('utf-8-sig')
                st.download_button(
                    "📥 比較結果をCSVダウンロード",
                    csv,
                    "forecast_vs_actual_comparison.csv",
                    "text/csv",
                    key='download_comparison'
                )
            else:
                st.warning("比較するデータがありません。")
        
        
        elif st.session_state.page == "シナリオ比較":
            st.title("シナリオ比較")
            
            st.markdown("""
            <div class="info-box">
                <strong>💡 概要:</strong> 3つのシナリオ（現実・楽観・悲観）を横並びで比較します。
            </div>
            """, unsafe_allow_html=True)
            
            # 3シナリオのデータを取得
            scenarios = ["現実", "楽観", "悲観"]
            scenario_data = {}
            
            for scenario in scenarios:
                forecast_data = load_forecast_data_cached(
                    st.session_state.selected_period_id,
                    scenario,
                    processor
                )
                
                if not forecast_data.empty:
                    # 合計を計算
                    month_cols = [col for col in forecast_data.columns if col not in ['項目名']]
                    scenario_data[scenario] = {}
                    
                    for item in processor.all_items:
                        item_row = forecast_data[forecast_data['項目名'] == item]
                        if not item_row.empty:
                            total = 0
                            for month in month_cols:
                                if month in item_row.columns:
                                    val = item_row[month].iloc[0]
                                    if pd.notna(val):
                                        total += float(val)
                            scenario_data[scenario][item] = total
            
            if scenario_data:
                # 比較テーブル
                st.markdown("### 📊 シナリオ別 損益比較（通期）")
                
                comparison_rows = []
                
                # 主要項目のみ表示
                key_items = [
                    "売上高", "売上原価", "販売費及び一般管理費合計",
                    "営業損益金額", "経常損益金額", "当期純損益金額"
                ]
                
                for item in key_items:
                    if item in scenario_data.get("現実", {}):
                        row = {"項目": item}
                        
                        base_value = scenario_data["現実"].get(item, 0)
                        row["現実"] = base_value
                        
                        for scenario in ["楽観", "悲観"]:
                            value = scenario_data[scenario].get(item, 0)
                            row[scenario] = value
                            
                            # 差異率を計算
                            if base_value != 0:
                                diff_pct = ((value - base_value) / base_value * 100)
                                row[f"{scenario}_diff"] = f"{diff_pct:+.1f}%"
                            else:
                                row[f"{scenario}_diff"] = "-"
                        
                        comparison_rows.append(row)
                
                # 利益率を追加
                if "売上高" in scenario_data.get("現実", {}):
                    for profit_item, label in [
                        ("営業損益金額", "営業利益率"),
                        ("経常損益金額", "経常利益率"),
                        ("当期純損益金額", "純利益率")
                    ]:
                        row = {"項目": label}
                        
                        for scenario in scenarios:
                            sales = scenario_data[scenario].get("売上高", 0)
                            profit = scenario_data[scenario].get(profit_item, 0)
                            
                            if sales > 0:
                                rate = (profit / sales * 100)
                                row[scenario] = rate
                            else:
                                row[scenario] = 0
                        
                        # 差異ポイント
                        base_rate = row.get("現実", 0)
                        for scenario in ["楽観", "悲観"]:
                            rate = row.get(scenario, 0)
                            diff_pt = rate - base_rate
                            row[f"{scenario}_diff"] = f"{diff_pt:+.1f}pt"
                        
                        comparison_rows.append(row)
                
                comparison_df = pd.DataFrame(comparison_rows)
                
                # フォーマット設定
                def format_row(row):
                    if "率" in row['項目']:
                        # 利益率の行
                        return {
                            '項目': row['項目'],
                            '現実': f"{row['現実']:.1f}%",
                            '楽観': f"{row['楽観']:.1f}%",
                            '楽観_差異': row['楽観_diff'],
                            '悲観': f"{row['悲観']:.1f}%",
                            '悲観_差異': row['悲観_diff']
                        }
                    else:
                        # 金額の行
                        return {
                            '項目': row['項目'],
                            '現実': f"¥{safe_int(row['現実']):,}",
                            '楽観': f"¥{safe_int(row['楽観']):,}",
                            '楽観_差異': row['楽観_diff'],
                            '悲観': f"¥{safe_int(row['悲観']):,}",
                            '悲観_差異': row['悲観_diff']
                        }
                
                formatted_rows = [format_row(row) for _, row in comparison_df.iterrows()]
                display_df = pd.DataFrame(formatted_rows)
                
                # カラム名を整理
                display_df.columns = ['項目', '現実', '楽観', '差異', '悲観', '差異 ']
                
                st.dataframe(
                    display_df,
                    width="stretch",
                    height=500,
                    hide_index=True
                )
                
                st.markdown("---")
                
                # グラフ: 営業利益の比較
                st.markdown("### 📈 営業利益の比較")
                
                operating_profits = []
                for scenario in scenarios:
                    operating_profits.append(scenario_data[scenario].get("営業損益金額", 0))
                
                fig = go.Figure(data=[
                    go.Bar(
                        x=scenarios,
                        y=operating_profits,
                        marker_color=['#1976d2', '#2e7d32', '#f57c00'],
                        text=[f"¥{safe_int(v):,}" for v in operating_profits],
                        textposition='auto',
                    )
                ])
                
                fig.update_layout(
                    yaxis_title="営業利益（円）",
                    height=400,
                    template="plotly_white"
                )
                
                st.plotly_chart(fig, width="stretch")
                
                # CSVダウンロード
                st.markdown("---")
                csv = display_df.to_csv(index=False).encode('utf-8-sig')
                st.download_button(
                    label="📥 CSV形式でダウンロード",
                    data=csv,
                    file_name=f"scenario_comparison_{st.session_state.selected_period_id}.csv",
                    mime="text/csv"
                )
            else:
                st.warning("シナリオデータがありません。")
        
        elif st.session_state.page == "期間比較分析":
            st.title("期間比較")
            
            st.markdown("""
            <div class="info-box">
                <strong>💡 使い方:</strong> 異なる会計期間のデータを比較します。前期比、成長率などを確認できます。
            </div>
            """, unsafe_allow_html=True)
            
            # 期間選択
            all_periods = get_company_periods_cached(st.session_state.selected_comp_id, processor)
            
            if len(all_periods) < 2:
                st.warning("比較するには2期以上のデータが必要です。")
            else:
                col1, col2 = st.columns(2)
                
                with col1:
                    period1_id = st.selectbox(
                        "比較元の期",
                        all_periods['id'].tolist(),
                        format_func=lambda x: f"第{all_periods[all_periods['id']==x]['period_num'].iloc[0]}期",
                        key="period1"
                    )
                
                with col2:
                    period2_id = st.selectbox(
                        "比較先の期",
                        all_periods['id'].tolist(),
                        format_func=lambda x: f"第{all_periods[all_periods['id']==x]['period_num'].iloc[0]}期",
                        index=1 if len(all_periods) > 1 else 0,
                        key="period2"
                    )
                
                if period1_id == period2_id:
                    st.warning("異なる期を選択してください。")
                else:
                    # 両期間のデータを取得
                    data1 = load_actual_data_cached(period1_id, processor)
                    data2 = load_actual_data_cached(period2_id, processor)
                    
                    months1 = get_fiscal_months_cached(st.session_state.selected_comp_id, period1_id, processor)
                    months2 = get_fiscal_months_cached(st.session_state.selected_comp_id, period2_id, processor)
                    
                    # 比較テーブルを作成
                    period_comparison_rows = []
                    
                    for item in processor.all_items:
                        row1 = data1[data1['項目名'] == item]
                        row2 = data2[data2['項目名'] == item]
                        
                        if row1.empty or row2.empty:
                            continue
                        
                        # 合計計算
                        total1 = sum([float(row1[m].iloc[0]) if m in row1.columns and pd.notna(row1[m].iloc[0]) else 0 for m in months1])
                        total2 = sum([float(row2[m].iloc[0]) if m in row2.columns and pd.notna(row2[m].iloc[0]) else 0 for m in months2])
                        
                        diff = total2 - total1
                        growth_rate = (diff / total1 * 100) if total1 != 0 else 0
                        
                        period_comparison_rows.append({
                            "項目名": item,
                            f"第{all_periods[all_periods['id']==period1_id]['period_num'].iloc[0]}期": total1,
                            f"第{all_periods[all_periods['id']==period2_id]['period_num'].iloc[0]}期": total2,
                            "増減額": diff,
                            "成長率(%)": growth_rate
                        })
                    
                    period_comparison_df = pd.DataFrame(period_comparison_rows)
                    
                    if not period_comparison_df.empty:
                        formatted_period_df = period_comparison_df.style\
                            .format({
                                f"第{all_periods[all_periods['id']==period1_id]['period_num'].iloc[0]}期": "¥{:,.0f}",
                                f"第{all_periods[all_periods['id']==period2_id]['period_num'].iloc[0]}期": "¥{:,.0f}",
                                "増減額": "¥{:,.0f}",
                                "成長率(%)": "{:.1f}%"
                            })\
                            .applymap(
                                lambda x: 'background-color: #d4edda' if isinstance(x, (int, float)) and x > 0 else 
                                          ('background-color: #f8d7da' if isinstance(x, (int, float)) and x < 0 else ''),
                                subset=['増減額', '成長率(%)']
                            )
                        
                        st.dataframe(formatted_period_df, width="stretch", height=600)
                        
                        # CSVダウンロード
                        csv = period_comparison_df.to_csv(index=False).encode('utf-8-sig')
                        st.download_button(
                            "📥 期間比較結果をCSVダウンロード",
                            csv,
                            "period_comparison.csv",
                            "text/csv",
                            key='download_period_comparison'
                        )
                    else:
                        st.warning("比較するデータがありません。")
        
        elif st.session_state.page == "データインポート":
            st.title("データ取込")
            
            # タブで実績データと予測データを分ける
            tab1, tab2 = st.tabs(["💰 実績データインポート（BS・PL統合）", "📊 予測データインポート"])
            
            # ===== タブ1: 実績データインポート（BS・PL統合） =====
            with tab1:
                st.markdown("### 📊 BS・PLを含むExcelファイルから実績データを一括取り込み")
                st.caption("弥生会計からエクスポートしたExcelファイルをアップロード")
                
                st.info("""
                💡 **対応ファイル形式:**
                - シート「貸･事業所(合計)」（BS: 貸借対照表）
                - シート「損･事業所(合計)」（PL: 損益計算書）
                
                ファイルから自動で以下を認識します:
                - 会社名（A3セル）
                - 処理日時（A4セル）※最新データかチェック
                - 会計期間（A5セル）
                """)
                
                uploaded_file = st.file_uploader(
                    "Excelファイルを選択（BS・PL両シート含む）",
                    type=['xlsx', 'xls'],
                    help="弥生会計からエクスポートしたBS・PLファイルをアップロード",
                    key="actual_bs_pl_upload"
                )
                
                # ファイルが削除された場合のキャッシュクリア
                if uploaded_file is None:
                    for key in ['bs_data', 'pl_data', 'cf_data', 'file_metadata', 'show_actual_import']:
                        if key in st.session_state:
                            del st.session_state[key]
                
                if uploaded_file:
                    if 'bs_data' not in st.session_state:
                        with st.spinner("📂 BS・PLファイルを読み込んでいます..."):
                            try:
                                # 一時ファイルに保存
                                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
                                    tmp_file.write(uploaded_file.read())
                                    temp_path = tmp_file.name
                                
                                # Excelファイルを開く
                                import openpyxl
                                wb = openpyxl.load_workbook(temp_path, read_only=True, data_only=True)
                                
                                # シート名を確認
                                has_bs = '貸･事業所(合計)' in wb.sheetnames
                                has_pl = '損･事業所(合計)' in wb.sheetnames
                                
                                if not has_bs or not has_pl:
                                    st.error("❌ 必要なシートが見つかりません")
                                    st.markdown(f"""
                                    **検出されたシート:** {', '.join(wb.sheetnames)}
                                    
                                    **必要なシート:**
                                    - 貸･事業所(合計) {'✅' if has_bs else '❌'}
                                    - 損･事業所(合計) {'✅' if has_pl else '❌'}
                                    """)
                                else:
                                    # メタデータを読み込み（BSシートから）
                                    import openpyxl
                                    wb_meta = openpyxl.load_workbook(temp_path, read_only=True, data_only=True)
                                    sheet_meta = wb_meta['貸･事業所(合計)']
                                    
                                    # A3: 会社名
                                    company_name_raw = str(sheet_meta['A3'].value) if sheet_meta['A3'].value else ""
                                    company_name = company_name_raw.replace('事業所名：', '').strip()
                                    
                                    # A4: 処理日時
                                    process_time_raw = str(sheet_meta['A4'].value) if sheet_meta['A4'].value else ""
                                    process_time = process_time_raw.replace('処理日時：', '').strip()
                                    
                                    # A5: 会計期間
                                    period_raw = str(sheet_meta['A5'].value) if sheet_meta['A5'].value else ""
                                    period_parts = period_raw.replace('集計期間：', '').strip().split(',')
                                    
                                    period_start = period_parts[0] if len(period_parts) > 0 else ""
                                    period_end = period_parts[1] if len(period_parts) > 1 else ""
                                    
                                    wb_meta.close()
                                    
                                    # メタデータを保存
                                    metadata = {
                                        'company_name': company_name,
                                        'process_time': process_time,
                                        'period_start': period_start,
                                        'period_end': period_end
                                    }
                                    st.session_state.file_metadata = metadata
                                    
                                    # BSを読み込み
                                    bs_data = cf_analyzer.load_bs_from_yayoi(temp_path, sheet_name='貸･事業所(合計)')
                                    
                                    # PLを読み込み（既存のimport_yayoi_excelを使用）
                                    pl_df, info = processor.import_yayoi_excel(
                                        temp_path,
                                        st.session_state.selected_period_id,
                                        preview_only=True
                                    )
                                    
                                    # 一時ファイル削除
                                    if os.path.exists(temp_path):
                                        os.unlink(temp_path)
                                    
                                    if not bs_data.empty and pl_df is not None and not pl_df.empty:
                                        st.session_state.bs_data = bs_data
                                        st.session_state.pl_data = pl_df
                                        st.session_state.show_actual_import = True
                                        
                                        # CFを自動計算
                                        with st.spinner("💰 キャッシュフローを計算しています..."):
                                            cf_data = cf_analyzer.calculate_cash_flow(
                                                pl_df,
                                                bs_data
                                            )
                                            
                                            if cf_data:
                                                st.session_state.cf_data = cf_data
                                        
                                        st.success(f"✅ ファイル **{uploaded_file.name}** を読み込みました")
                                    else:
                                        st.error("❌ BS・PLの読み込みに失敗しました")
                            
                            except Exception as e:
                                st.error(f"❌ エラー: {str(e)}")
                                import traceback
                                st.code(traceback.format_exc())
                    
                    # データが読み込まれている場合
                    if st.session_state.get('show_actual_import'):
                        metadata = st.session_state.file_metadata
                        
                        st.markdown("---")
                        st.markdown("### 📋 ファイル情報")
                        
                        # ファイルメタデータを表示
                        col1, col2 = st.columns(2)
                        with col1:
                            st.markdown(f"""
                            **🏢 会社名:** {metadata['company_name']}  
                            **📅 会計期間:** {metadata['period_start']} 〜 {metadata['period_end']}
                            """)
                        with col2:
                            st.markdown(f"""
                            **⏰ 処理日時:** {metadata['process_time']}
                            """)
                        
                        # 会社名チェック
                        if metadata['company_name'] != st.session_state.selected_comp_name:
                            st.warning(f"⚠️ ファイルの会社名（{metadata['company_name']}）と選択中の会社名（{st.session_state.selected_comp_name}）が異なります")
                        
                        st.markdown("---")
                        st.markdown("### 📊 データサマリー")
                        
                        col1, col2, col3, col4 = st.columns(4)
                        with col1:
                            st.metric("BS項目数", len(st.session_state.bs_data))
                        with col2:
                            st.metric("PL項目数", len(st.session_state.pl_data))
                        with col3:
                            month_cols = [c for c in st.session_state.bs_data.columns if '月度' in str(c)]
                            st.metric("月数", len(month_cols))
                        with col4:
                            if 'cf_data' in st.session_state:
                                st.metric("CF計算", "✅ 完了")
                            else:
                                st.metric("CF計算", "❌ 未完了")
                        
                        # タブでBS・PL・CFを表示
                        data_tab1, data_tab2, data_tab3 = st.tabs(["📈 PL（損益計算書）", "📊 BS（貸借対照表）", "💰 CF（キャッシュフロー）"])
                        
                        with data_tab1:
                            st.markdown("#### PLデータ プレビュー")
                            st.dataframe(
                                st.session_state.pl_data.head(20),
                                width="stretch",
                                height=400
                            )
                        
                        with data_tab2:
                            st.markdown("#### BSデータ プレビュー")
                            st.dataframe(
                                st.session_state.bs_data.head(20),
                                width="stretch",
                                height=400
                            )
                        
                        with data_tab3:
                            if 'cf_data' in st.session_state and st.session_state.cf_data:
                                st.markdown("#### CFデータ プレビュー")
                                cf_rows = []
                                for month, cf_month_data in st.session_state.cf_data.items():
                                    if '営業CF' in cf_month_data and '合計' in cf_month_data['営業CF']:
                                        cf_rows.append({
                                            '月': month,
                                            '営業CF': cf_month_data['営業CF']['合計'],
                                            '投資CF': cf_month_data['投資CF'].get('合計', 0),
                                            '財務CF': cf_month_data['財務CF'].get('合計', 0),
                                            '現金増減': cf_month_data.get('現金増減', 0),
                                            '期末現金': cf_month_data.get('期末現金', 0)
                                        })
                                
                                if cf_rows:
                                    cf_df = pd.DataFrame(cf_rows)
                                    formatted_cf = cf_df.style.format({
                                        '営業CF': '¥{:,.0f}',
                                        '投資CF': '¥{:,.0f}',
                                        '財務CF': '¥{:,.0f}',
                                        '現金増減': '¥{:,.0f}',
                                        '期末現金': '¥{:,.0f}'
                                    })
                                    st.dataframe(formatted_cf, width="stretch", height=400)
                            else:
                                st.info("CF計算データがありません")
                        
                        st.markdown("---")
                        
                        # インポートボタン
                        st.warning("⚠️ インポートを実行すると、以下のデータがデータベースに保存されます")
                        st.markdown("""
                        - ✅ PL（損益計算書）実績データ
                        - ✅ BS（貸借対照表）データ
                        - ✅ CF（キャッシュフロー）データ
                        """)
                        
                        col1, col2 = st.columns([1, 3])
                        with col1:
                            if st.button("✅ インポート実行", type="primary", width="stretch"):
                                progress_bar = st.progress(0)
                                status_text = st.empty()
                                
                                try:
                                    # PL実績データ保存
                                    status_text.text("💾 PLデータを保存中...")
                                    progress_bar.progress(20)
                                    
                                    success_pl, info_pl = processor.save_extracted_data(
                                        st.session_state.selected_period_id,
                                        st.session_state.pl_data
                                    )
                                    
                                    # BSデータ保存
                                    status_text.text("💾 BSデータを保存中...")
                                    progress_bar.progress(40)
                                    
                                    success_bs = cf_analyzer.save_bs_to_db(
                                        st.session_state.selected_period_id,
                                        st.session_state.bs_data
                                    )
                                    
                                    # CFデータ保存
                                    status_text.text("💾 CFデータを保存中...")
                                    progress_bar.progress(60)
                                    
                                    success_cf = False
                                    if 'cf_data' in st.session_state and st.session_state.cf_data:
                                        success_cf = cf_analyzer.save_cf_to_db(
                                            st.session_state.selected_period_id,
                                            st.session_state.cf_data
                                        )
                                    
                                    progress_bar.progress(80)
                                    
                                    # 結果確認
                                    if success_pl and success_bs:
                                        progress_bar.progress(100)
                                        status_text.text("✅ 完了")
                                        st.success("✅ インポートが完了しました！")
                                        
                                        if success_cf:
                                            st.success("✅ CFデータも保存されました")
                                        
                                        # キャッシュクリア
                                        for key in ['bs_data', 'pl_data', 'cf_data', 'file_metadata', 'show_actual_import', 
                                                   'actuals_df', 'imported_df', 'pl_df', 'forecast_data_cache']:
                                            if key in st.session_state:
                                                del st.session_state[key]
                                        
                                        st.cache_data.clear()
                                        
                                        # リロード
                                        import time
                                        time.sleep(1)
                                        st.rerun()
                                    else:
                                        progress_bar.empty()
                                        status_text.empty()
                                        st.error(f"❌ インポートに失敗しました")
                                        if not success_pl:
                                            st.error(f"PL保存エラー: {info_pl}")
                                        if not success_bs:
                                            st.error("BS保存エラー")
                                
                                except Exception as e:
                                    progress_bar.empty()
                                    status_text.empty()
                                    st.error(f"❌ エラーが発生しました: {str(e)}")
                                    import traceback
                                    st.code(traceback.format_exc())
                        
                        with col2:
                            if st.button("🔄 キャンセル", width="stretch"):
                                # キャッシュクリア
                                for key in ['bs_data', 'pl_data', 'cf_data', 'file_metadata', 'show_actual_import']:
                                    if key in st.session_state:
                                        del st.session_state[key]
                                st.rerun()
            
            # ===== タブ2: 予測データインポート =====
            with tab2:
                st.markdown("""
                <div class="info-box">
                    <strong>💡 使い方:</strong><br>
                    1. テンプレートをダウンロード<br>
                    2. Excelで予測数値を入力<br>
                    3. ファイルをアップロード
                </div>
                """, unsafe_allow_html=True)
                
                # シナリオ選択
                forecast_scenario = st.selectbox(
                    "インポート先シナリオを選択",
                    ["現実", "楽観", "悲観"],
                    key="forecast_import_scenario"
                )
                
                # テンプレートダウンロード
                st.subheader("📥 ステップ1: テンプレートをダウンロード")
                
                template_df = processor.create_forecast_template(
                    st.session_state.selected_period_id,
                    forecast_scenario
                )
                
                if template_df is not None:
                    # Excelファイルとして出力
                    from io import BytesIO
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        template_df.to_excel(writer, index=False, sheet_name='予測データ')
                    excel_data = output.getvalue()
                    
                    st.download_button(
                        label="📥 予測データテンプレートをダウンロード",
                        data=excel_data,
                        file_name=f"予測データテンプレート_{forecast_scenario}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary"
                    )
                    
                    st.info("""
                    💡 **テンプレートの使い方:**
                    - 各項目の予測数値を月ごとに入力してください
                    - 0のままの項目はインポートされません
                    - 項目名の列は変更しないでください
                    """)
                
                st.markdown("---")
                
                # ファイルアップロード
                st.subheader("📤 ステップ2: 入力済みファイルをアップロード")
                
                forecast_file = st.file_uploader(
                    "予測データExcelファイルを選択",
                    type=['xlsx', 'xls'],
                    help="入力済みのテンプレートファイルをアップロードしてください",
                    key="forecast_upload"
                )
                
                # ファイルが削除された場合のキャッシュクリア
                if forecast_file is None:
                    if 'forecast_imported_df' in st.session_state:
                        del st.session_state.forecast_imported_df
                    if 'show_forecast_import_button' in st.session_state:
                        del st.session_state.show_forecast_import_button
                
                if forecast_file:
                    if 'forecast_imported_df' not in st.session_state:
                        try:
                            # Excelファイルを読み込み（シート名を指定）
                            try:
                                # まず「予測データ」シートを試す
                                forecast_df = pd.read_excel(forecast_file, sheet_name='予測データ')
                            except:
                                # 失敗したら最初のシートを読み込む
                                forecast_df = pd.read_excel(forecast_file, sheet_name=0)
                            
                            # 基本的なバリデーション
                            if '項目名' not in forecast_df.columns:
                                st.error("❌ テンプレート形式が正しくありません。「項目名」列が見つかりません。")
                                st.info("""
                                💡 **確認事項:**
                                - テンプレートファイルを使用していますか？
                                - 「項目名」列を削除していませんか？
                                - シート名は「予測データ」ですか？
                                """)
                            else:
                                st.success(f"✅ ファイル **{forecast_file.name}** を読み込みました")
                                st.session_state.forecast_imported_df = forecast_df
                                st.session_state.show_forecast_import_button = True
                        
                        except Exception as e:
                            st.error(f"❌ ファイルの読み込みに失敗しました: {str(e)}")
                            st.info("""
                            💡 **よくある原因:**
                            - ファイルが破損している
                            - Excel形式ではない
                            - テンプレート形式と異なる
                            """)
                    
                    if st.session_state.get('show_forecast_import_button'):
                        st.subheader("📋 インポートデータ プレビュー（直接編集可能）")
                        
                        st.markdown("""
                        <div class="info-box">
                            <strong>✏️ 編集:</strong> セルをダブルクリックして値を直接修正できます。
                        </div>
                        """, unsafe_allow_html=True)
                        
                        # 編集可能なデータエディタを使用
                        edited_forecast_df = st.data_editor(
                            st.session_state.forecast_imported_df,
                            width="stretch",
                            height=400,
                            num_rows="fixed",
                            disabled=["項目名"],
                            column_config={
                                col: st.column_config.NumberColumn(
                                    format="%.0f",
                                    min_value=-999999999,
                                    max_value=999999999
                                ) for col in st.session_state.forecast_imported_df.columns if col != '項目名'
                            }
                        )
                        
                        # 編集後のデータを保存
                        st.session_state.forecast_imported_df = edited_forecast_df
                        
                        st.markdown(f"""
                        <div class="warning-box">
                            <strong>⚠️ 注意:</strong> 上記の内容でインポートを実行すると、「{forecast_scenario}」シナリオの予測データが上書きされます。
                        </div>
                        """, unsafe_allow_html=True)
                        
                        if st.button("✅ 予測データをインポート", type="primary", key="import_forecast"):
                            success, info = processor.save_forecast_from_excel(
                                st.session_state.selected_period_id,
                                forecast_scenario,
                                st.session_state.forecast_imported_df
                            )
                            if success:
                                st.success(f"✅ {info}")
                                # 全キャッシュをクリア
                                st.cache_data.clear()
                                # すべての関連キャッシュを削除
                                for key in ['forecasts_df', 'forecast_imported_df', 'show_forecast_import_button',
                                           'pl_df', 'sub_accounts_df', 'actuals_df',
                                           'scenario_adjustment_cache', 'sub_account_aggregation_cache',
                                           'forecast_input_cache_key', 'forecast_input_data',
                                           'adjustment_key', 'sub_cache_key', 'pl_cache_key']:
                                    if key in st.session_state:
                                        del st.session_state[key]
                                st.rerun()
                            else:
                                st.error(f"❌ インポートに失敗しました: {info}")
            
            
        
        elif st.session_state.page == "収益構造分析":
            show_profitability_analysis_page(processor)
        
        elif st.session_state.page == "シナリオ一括設定":
            st.title("シナリオ一括設定")
            
            st.markdown("""
            <div class="info-box">
                <strong>💡 使い方:</strong> 「現実」シナリオをベースに、「楽観」「悲観」シナリオの増減率を設定します。
                設定した増減率は全画面に即座に反映されます。
            </div>
            """, unsafe_allow_html=True)
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("### 📈 楽観シナリオ")
                st.markdown("""
                <div class="success-box">
                    <strong>想定される効果:</strong><br>
                    • 売上: 増加率そのまま適用<br>
                    • 売上原価: 増加率の50%を逆方向に適用<br>
                    • 販管費: 増加率の30%を逆方向に適用
                </div>
                """, unsafe_allow_html=True)
                
                new_opt_rate = st.number_input(
                    "楽観シナリオ増減率 (%)",
                    value=st.session_state.scenario_rates["楽観"] * 100,
                    min_value=-100.0,
                    max_value=100.0,
                    step=1.0,
                    key="opt_rate_input"
                ) / 100.0
                
                if st.button("💾 楽観シナリオ増減率を保存", type="primary"):
                    st.session_state.scenario_rates["楽観"] = new_opt_rate
                    st.success(f"✅ 楽観シナリオの増減率を **{new_opt_rate * 100:.1f}%** に設定しました")
                    st.rerun()
            
            with col2:
                st.markdown("### 📉 悲観シナリオ")
                st.markdown("""
                <div class="warning-box">
                    <strong>想定される効果:</strong><br>
                    • 売上: 減少率そのまま適用<br>
                    • 売上原価: 減少率の50%を逆方向に適用<br>
                    • 販管費: 減少率の30%を逆方向に適用
                </div>
                """, unsafe_allow_html=True)
                
                new_pes_rate = st.number_input(
                    "悲観シナリオ増減率 (%)",
                    value=st.session_state.scenario_rates["悲観"] * 100,
                    min_value=-100.0,
                    max_value=100.0,
                    step=1.0,
                    key="pes_rate_input"
                ) / 100.0
                
                if st.button("💾 悲観シナリオ増減率を保存", type="primary"):
                    st.session_state.scenario_rates["悲観"] = new_pes_rate
                    st.success(f"✅ 悲観シナリオの増減率を **{new_pes_rate * 100:.1f}%** に設定しました")
                    st.rerun()
            
            st.markdown("---")
            
            # 設定値サマリー
            st.subheader("📋 現在の設定値")
            
            summary_data = {
                "シナリオ": ["現実", "楽観", "悲観"],
                "増減率": [
                    f"{st.session_state.scenario_rates['現実'] * 100:.1f}%",
                    f"{st.session_state.scenario_rates['楽観'] * 100:.1f}%",
                    f"{st.session_state.scenario_rates['悲観'] * 100:.1f}%"
                ],
                "説明": [
                    "ベースとなる予測値",
                    "売上増加・費用削減を想定",
                    "売上減少・費用増加を想定"
                ]
            }
            
            st.table(pd.DataFrame(summary_data))
        

# AI自動予測ページ
elif st.session_state.page == "AI自動予測":
    st.title("🔮 AI自動予測")
    
    # デバッグ情報を常に表示
    with st.expander("🔍 デバッグ情報", expanded=False):
        st.write("**ADVANCED_FORECAST_AVAILABLE:**", ADVANCED_FORECAST_AVAILABLE)
        st.write("**selected_period_id:**", st.session_state.get('selected_period_id'))
        import os
        py_files = [f for f in os.listdir('.') if f.endswith('.py')]
        st.write("**Python files:**", py_files[:10])
    
    if not ADVANCED_FORECAST_AVAILABLE:
        st.error("❌ AI予測機能が利用できません")
        st.info("""
        **必要なファイル:**
        - advanced_forecast_engine.py
        - advanced_forecast_ui.py
        
        これらのファイルを配置してアプリを再起動してください。
        """)
        st.stop()
    
    try:
        # データハンドラー
        if 'data_handler_adapter' not in st.session_state:
            class DataHandlerAdapter:
                def __init__(self, proc):
                    self._processor = proc
                
                def get_actual_vs_forecast_split(self, period_id):
                    try:
                        actuals = self._processor.load_actual_data(period_id)
                        if actuals is None or actuals.empty:
                            return {
                                'has_actual': False,
                                'latest_actual_month': 0,
                                'actual_months': [],
                                'forecast_months': list(range(1, 13))
                            }
                        
                        latest = int(actuals['fiscal_month'].max())
                        return {
                            'has_actual': True,
                            'latest_actual_month': latest,
                            'actual_months': list(range(1, latest + 1)),
                            'forecast_months': list(range(latest + 1, 13))
                        }
                    except Exception as e:
                        st.error(f"Split error: {e}")
                        return {
                            'has_actual': False,
                            'latest_actual_month': 0,
                            'actual_months': [],
                            'forecast_months': list(range(1, 13))
                        }
                
                def get_cumulative_actual_data(self, period_id, up_to_month):
                    try:
                        actuals = self._processor.load_actual_data(period_id)
                        if actuals is None or actuals.empty:
                            return {}
                        actuals = actuals[actuals['fiscal_month'] <= up_to_month]
                        result = {}
                        for _, row in actuals.iterrows():
                            account = row.get('account_name', '')
                            amount = row.get('amount', 0)
                            if account:
                                result[account] = amount
                        return result
                    except Exception as e:
                        st.error(f"Cumulative error: {e}")
                        return {}
            
            st.session_state.data_handler_adapter = DataHandlerAdapter(processor)
        
        # 予測エンジン
        if 'advanced_engine' not in st.session_state:
            st.session_state.advanced_engine = get_advanced_forecast_engine(
                st.session_state.data_handler_adapter
            )
        
        # 予測画面表示
        st.success("✅ AI予測機能が正常にロードされました")
        show_advanced_forecast_page(
            st.session_state.data_handler_adapter,
            st.session_state.advanced_engine
        )
    
    except Exception as e:
        st.error(f"❌ エラー: {e}")
        with st.expander("詳細", expanded=True):
            import traceback
            st.code(traceback.format_exc())

else:
    # 会社または期が未登録の場合
    if companies.empty:
        st.title("財務予測シミュレーター")
        
        st.markdown("""
        <div style="background-color: #e3f2fd; padding: 2rem; border-radius: 10px; margin: 2rem 0;">
            <h3 style="color: #1976d2; margin-top: 0;">🚀 はじめての方へ</h3>
            <p style="font-size: 1.1rem; line-height: 1.8;">
                まずは以下の手順でセットアップしてください：
            </p>
            <div style="background-color: white; padding: 1.5rem; border-radius: 8px; margin: 1rem 0;">
                <strong style="font-size: 1.2rem; color: #1976d2;">📍 手順</strong><br><br>
                <strong style="color: #d32f2f;">1️⃣ 左サイドバーの「⚙️ システム設定」をクリック</strong><br>
                <span style="font-size: 0.9rem; color: #666;">← 左側のメニューから選択してください</span><br><br>
                <strong>2️⃣ 会社設定タブで会社名を入力</strong><br><br>
                <strong>3️⃣ 会計期間設定タブで期の情報を入力</strong><br><br>
                <strong>4️⃣ サイドバーで会社と期を選択</strong><br>
                <span style="font-size: 0.9rem; color: #666;">→ すべての機能が使えるようになります！</span>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        # データベース接続状態を表示
        if processor.use_postgres:
            st.success("✅ Supabaseに接続済み - データは永続的に保存されます")
        else:
            st.info("ℹ️ ローカルモードで動作中")
            
    else:
        st.warning("### ⚠️ 会計期間が選択されていません")
        st.markdown("""
        <div class="warning-box">
            <strong>会計期間を登録してください</strong><br><br>
            左サイドバーの「システム設定」→「会計期間設定」タブから<br>
            会計期間を追加してください。
        </div>
        """, unsafe_allow_html=True)
